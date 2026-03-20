from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse
from zoneinfo import ZoneInfo
from django.http import Http404, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, render, HttpResponse
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, View, DeleteView, TemplateView, FormView
from django.db.models import Q, Count, ExpressionWrapper, Sum, Max, Value, CharField, F, Case, When
from django.db.models.fields import DecimalField, IntegerField
from django.shortcuts import redirect
from httpcore import request
from sqlalchemy import Cast
from .models import BajaStock, Ingreso, IngresoItem, Local, Marca, MovimientoStock, Producto, Articulo, ProductoBulkAdjust, ProductoBulkAdjustItem, Promocion, RetiroCaja, Transferencia, TransferenciaItem, Venta, VentaItem, VentaArticulo
from .forms import ArticuloEditForm, ArticuloImportXlsxForm, CheckoutForm, ProductoImportXlsxForm, PromocionForm, TransferirArticuloForm, UserLoginForm, UserRegisterForm, ArticuloCreateForm, ArticuloImportXlsxForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic.edit import FormView
from django.db import transaction
from django.contrib import messages
from datetime import datetime as Datetime, time, timedelta, timezone, datetime
from django.db.models.functions import TruncDate, Coalesce, TruncMinute, Concat
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from django.utils import timezone
from openpyxl import load_workbook
from django.core.mail import EmailMessage



@login_required
def index(request):
    local = _get_local_activo(request)
    hoy = timezone.localdate()

    ventas_qs = (
        Venta.objects
        .filter(
            usuario=request.user,
            local=local if local else None,
            estado=Venta.Estado.CERRADA,
            fecha__date=hoy,
        )
        .order_by("-venta_id")
    )

    agg = ventas_qs.aggregate(
        ventas_hoy_count=Count("venta_id"),
        ventas_hoy_total=Coalesce(Sum("total"), Decimal("0.00")),
    )

    ventas_hoy = list(
        ventas_qs.values("venta_id", "fecha", "metodo_de_pago", "total")[:8]
    )

    caja_qs = (
        RetiroCaja.objects
        .filter(
            local=local if local else None,
            fecha=hoy,
        )
        .select_related("usuario")
        .order_by("-fecha")
    )

    if not request.user.is_staff:
        caja_qs = caja_qs.filter(usuario=request.user)

    movimientos_caja = list(caja_qs[:8])

    promociones_activas = Promocion.objects.filter(estado="ACT").count()
    total_en_efectivo = _saldo_caja_local(local) if local else Decimal("0.00")

    context = {
        "now": timezone.now(),
        "local_activo_id": getattr(local, "local_id", None),

        "ventas_hoy_count": agg["ventas_hoy_count"] or 0,
        "ventas_hoy_total": agg["ventas_hoy_total"] or Decimal("0.00"),
        "efectivo_en_caja": total_en_efectivo,
        "promociones_activas": promociones_activas,

        "ventas_hoy": ventas_hoy,
        "movimientos_caja": movimientos_caja,
    }
    return render(request, "inventory/index.html", context)

def _is_admin(user):
    return user.is_staff or user.is_superuser


class SignUpView(View):
    def get(self, request):
        form = UserRegisterForm()
        return render(request, "inventory/auth/signup.html", {"form": form})

    def post(self, request):
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            user = authenticate(username=form.cleaned_data['username'],
                                password=form.cleaned_data['password1'])
            login(request, user)
            return redirect("inventory:producto_list")
        return render(request, "inventory/auth/signup.html", {"form": form})

class LoginView(View):
    def get(self, request):
        if request.user.is_authenticated:
            return redirect("inventory:producto_list")
        form = UserLoginForm()
        return render(request, "inventory/auth/login.html", {"form": form})
    
    def post(self, request):
        form = UserLoginForm(request.POST)
        if form.is_valid():
            user = authenticate(
                request, 
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password'])
            if user is not None:
                login(request, user)
                return redirect("inventory:producto_list")
            else:
                form.add_error(None, "Credenciales inválidas")
        return render(request, "inventory/auth/login.html", {"form": form})
    
class LogoutView(View):
    def get(self, request):
        logout(request)
        return redirect("inventory:login")


class StaffRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff
    
class MarcaListView(LoginRequiredMixin, StaffRequiredMixin, ListView):
    model = Marca
    template_name = "inventory/marca/marca_list.html"
    context_object_name = "marcas"
    paginate_by = 30

    def get_queryset(self):
        q = (self.request.GET.get("q") or "").strip()
        sort = (self.request.GET.get("sort") or "nombre").strip().lower()

        qs = Marca.objects.all()

        if q:
            qs = qs.filter(nombre__icontains=q)

        qs = qs.annotate(productos_count=Count("productos", distinct=True))

        # Ventas GLOBAL: Marca -> Productos -> MovimientoStock (solo VENTA)
        sales_filter = Q(productos__movimientostock__tipo=MovimientoStock.Tipo.VENTA)

        qs = qs.annotate(
            unidades_raw=Coalesce(
                Sum("productos__movimientostock__cantidad", filter=sales_filter),
                Value(0),
            ),
            vendido_total=Coalesce(
                Sum("productos__movimientostock__precio_unitario", filter=sales_filter),
                Value(Decimal("0.00"), output_field=DecimalField(max_digits=12, decimal_places=2)),
            ),
            profit_total=Coalesce(
                Sum("productos__movimientostock__profit_unitario", filter=sales_filter),
                Value(Decimal("0.00"), output_field=DecimalField(max_digits=12, decimal_places=2)),
            ),
        ).annotate(
            # cantidad en ventas suele ser negativa -> la mostramos positiva
            unidades_vendidas=ExpressionWrapper(-F("unidades_raw"), output_field=IntegerField())
        )

        if sort == "ventas":
            qs = qs.order_by("-vendido_total", "nombre")
        elif sort == "productos":
            qs = qs.order_by("-productos_count", "nombre")
        else:
            qs = qs.order_by("nombre")

        return qs


    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = (self.request.GET.get("q") or "").strip()
        ctx["sort"] = (self.request.GET.get("sort") or "nombre").strip().lower()
        return ctx


class MarcaDetailView(LoginRequiredMixin, StaffRequiredMixin, DetailView):
    model = Marca
    template_name = "inventory/marca/marca_detail.html"
    context_object_name = "marca"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        marca = ctx["marca"]

        productos = Producto.objects.filter(marca=marca).order_by("nombre")
        ctx["productos"] = productos

        # Ventas GLOBAL de la marca
        sales_qs = MovimientoStock.objects.filter(
            producto__marca=marca,
            tipo=MovimientoStock.Tipo.VENTA,
        )

        agg = sales_qs.aggregate(
            vendido_total=Coalesce(
                Sum("precio_unitario"),
                Value(Decimal("0.00"), output_field=DecimalField(max_digits=12, decimal_places=2))
            ),
            profit_total=Coalesce(
                Sum("profit_unitario"),
                Value(Decimal("0.00"), output_field=DecimalField(max_digits=12, decimal_places=2))
            ),
            unidades_raw=Coalesce(Sum("cantidad"), Value(0)),
        )

        ctx["stats"] = {
            "productos_count": productos.count(),
            "vendido_total": agg["vendido_total"] or Decimal("0.00"),
            "profit_total": agg["profit_total"] or Decimal("0.00"),
            "unidades_vendidas": -int(agg["unidades_raw"] or 0),  # ventas negativas -> positivo
        }
        return ctx

class MarcaCreateView(LoginRequiredMixin, StaffRequiredMixin, CreateView):
    model = Marca
    fields = ["nombre"]
    template_name = "inventory/marca/marca_form.html"
    success_url = reverse_lazy("inventory:marca_list")


class MarcaUpdateView(LoginRequiredMixin, StaffRequiredMixin, UpdateView):
    model = Marca
    fields = ["nombre"]
    template_name = "inventory/marca/marca_form.html"

    def get_success_url(self):
        return reverse_lazy("inventory:marca_detail", kwargs={"pk": self.object.pk})


class MarcaDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
    model = Marca
    template_name = "inventory/marca/marca_confirm_delete.html"
    success_url = reverse_lazy("inventory:marca_list")
    
class ProductoListView(LoginRequiredMixin, ListView):
    model = Producto
    template_name = "inventory/producto/producto_list.html"
    context_object_name = "productos"
    paginate_by = 20  
    
    def get_queryset(self):
        qs = super().get_queryset().order_by("product_id")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(nombre__icontains=q)
        return qs
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = (self.request.GET.get("q") or "").strip()
        ctx["marcas"] = Marca.objects.all().order_by("nombre")
        return ctx  

class ProductoDetailView(DetailView):
    model = Producto
    template_name = "inventory/producto_list.html"
    context_object_name = "producto"
    
class ProductoCreateView(LoginRequiredMixin, StaffRequiredMixin,CreateView):
    model = Producto
    fields = ["nombre", "tipo_producto", "material", "marca", "precio", "costo"]
    template_name = "inventory/producto/producto_form.html"
    success_url = reverse_lazy("inventory:producto_list")


class ProductoUpdateView(LoginRequiredMixin, StaffRequiredMixin, UpdateView):
    model = Producto
    fields = ["nombre", "tipo_producto", "material", "marca", "precio", "costo"]
    template_name = "inventory/producto/producto_form.html"
    success_url = reverse_lazy("inventory:producto_list")


class ProductoDeleteView(LoginRequiredMixin,StaffRequiredMixin, View):
    model = Producto
    template_name = "inventory/producto/producto_confirm_delete.html"
    success_url = reverse_lazy("inventory:producto_list")
    
    def post(self, request, pk):
        producto = Producto.objects.get(pk=pk)
        producto.delete()
        return redirect(self.success_url)


class SetLocalView(LoginRequiredMixin, View):
    def post(self, request):
        local_id = request.POST.get("local_id")
        if Local.objects.filter(local_id=local_id).exists():
            request.session["local_id"] = int(local_id)
        return redirect(request.META.get("HTTP_REFERER", "/"))
    
def _should_show_all_locals(request):
    return _is_admin(request.user) and (request.GET.get("all_locals") == "1")

class ArticuloUpdateView(LoginRequiredMixin, StaffRequiredMixin, UpdateView):
    model = Articulo
    form_class = ArticuloEditForm
    context_object_name = "articulo"
    pk_url_kwarg = "articulo_id"
    template_name = "inventory/articulo/articulo_form.html"

    def get_success_url(self):
        return reverse_lazy("inventory:articulo_list")

class ArticulosBulkEditView(LoginRequiredMixin, StaffRequiredMixin, FormView):
    template_name = "inventory/articulo/articulo_bulk_form.html"
    form_class = ArticuloEditForm

    def dispatch(self, request, *args, **kwargs):
        local = _get_local_activo(request)
        if not local:
            messages.error(request, "No hay un local activo seleccionado.")
            return redirect("inventory:articulo_list")
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw["bulk"] = True
        return kw

    def _group(self):
        local = _get_local_activo(self.request)

        g = {
            "product_id": (self.request.GET.get("product_id") or self.request.POST.get("g_product_id") or "").strip(),
            "barcode": (self.request.GET.get("barcode") or self.request.POST.get("g_barcode") or "").strip(),
            "talle": (self.request.GET.get("talle") or self.request.POST.get("g_talle") or "").strip(),
            "color": (self.request.GET.get("color") or self.request.POST.get("g_color") or "").strip(),
            "sku": (self.request.GET.get("sku") or self.request.POST.get("g_sku") or "").strip(),
            "estado": (self.request.GET.get("estado") or self.request.POST.get("g_estado") or "DISP").strip().upper(),
        }

        if not all([g["product_id"], g["barcode"], g["talle"], g["color"], g["sku"], g["estado"]]):
            return None, None

        qs = Articulo.objects.filter(
            local=local,
            estado=g["estado"],
            product_id_id=g["product_id"],
            barcode=g["barcode"],
            talle=g["talle"],
            color=g["color"],
            sku=g["sku"],
        ).order_by("-articulo_id")

        return g, qs

    def get(self, request, *args, **kwargs):
        g, qs = self._group()
        if g is None:
            messages.error(request, "Faltan datos para editar por lote.")
            return redirect("inventory:articulo_list")
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        g, qs = self._group()
        ctx["group"] = g
        ctx["max_qty"] = qs.count() if qs is not None else 0
        return ctx

    @transaction.atomic
    def form_valid(self, form):
        g, qs = self._group()
        if g is None or qs is None:
            messages.error(self.request, "No se pudo resolver el grupo.")
            return redirect("inventory:articulo_list")

        if g["estado"] != "DISP":
            messages.error(self.request, "Solo podés editar por lote artículos DISP.")
            return redirect("inventory:articulo_list")

        all_flag = (self.request.POST.get("all") == "1")
        qty = int(self.request.POST.get("qty") or 0)
        total = qs.count()

        if total == 0:
            messages.warning(self.request, "No hay artículos para editar en ese grupo.")
            return redirect("inventory:articulo_list")

        if not all_flag:
            if qty <= 0:
                messages.error(self.request, "Cantidad inválida.")
                return redirect("inventory:articulo_list")
            qs = qs[:min(qty, total)]

        updates = {}
        for k in ["barcode", "product_id", "sku", "talle", "color"]:
            v = form.cleaned_data.get(k)
            if v not in (None, "", []):
                updates[k] = v

        ids = list(qs.values_list("articulo_id", flat=True))
        Articulo.objects.filter(articulo_id__in=ids).update(**updates)

        messages.success(self.request, f"Editados {len(ids)} artículo(s).")
        return redirect("inventory:articulo_list")
    
class ArticuloBajaView(LoginRequiredMixin, StaffRequiredMixin, View):
    @transaction.atomic
    def post(self, request, articulo_id):
        local = _get_local_activo(request)
        if not local:
            messages.error(request, "No hay un local activo seleccionado.")
            return redirect("inventory:articulo_list")

        a = get_object_or_404(Articulo, articulo_id=articulo_id, local=local)

        if a.estado != Articulo.Estado.DISPONIBLE:
            messages.error(request, "Solo podés dar de baja artículos DISP.")
            return redirect("inventory:articulo_list")

        a.estado = Articulo.Estado.BAJA
        a.save(update_fields=["estado"])
        baja = BajaStock.objects.create(usuario=request.user, local=local or "")
        MovimientoStock.objects.create(
            tipo=MovimientoStock.Tipo.BAJA,
            local=local,
            usuario=request.user,
            articulo=a,
            producto=a.product_id,
            sku=a.sku,
            barcode=a.barcode,
            talle=a.talle,
            color=a.color,
            cantidad=1,
            costo_unitario=Decimal(getattr(a.product_id, "costo", 0) or 0),
            precio_unitario=None,
            profit_unitario=Decimal("0.00"),
            baja=baja,
            nota="Baja manual",
        )

        messages.success(request, "Artículo dado de baja.")
        return redirect("inventory:articulo_list") 

class ArticulosBulkBajaView(LoginRequiredMixin, StaffRequiredMixin, View):
    @transaction.atomic
    def post(self, request):
        local = _get_local_activo(request)
        if not local:
            messages.error(request, "No hay un local activo seleccionado.")
            return redirect("inventory:articulo_list")

        g = {
            "product_id": (request.POST.get("g_product_id") or request.POST.get("product_id") or "").strip(),
            "barcode": (request.POST.get("g_barcode") or request.POST.get("barcode") or "").strip(),
            "talle": (request.POST.get("g_talle") or request.POST.get("talle") or "").strip(),
            "color": (request.POST.get("g_color") or request.POST.get("color") or "").strip(),
            "sku": (request.POST.get("g_sku") or request.POST.get("sku") or "").strip(),
            "estado": (request.POST.get("g_estado") or request.POST.get("estado") or "DISP").strip().upper(),
        }

        qs = Articulo.objects.filter(
            local=local,
            estado=Articulo.Estado.DISPONIBLE,
            product_id_id=g["product_id"],
            barcode=g["barcode"],
            talle=g["talle"],
            color=g["color"],
            sku=g["sku"],
        ).order_by("-articulo_id")

        total = qs.count()
        if total == 0:
            messages.warning(request, "No hay artículos para dar de baja.")
            return redirect("inventory:articulo_list")

        all_flag = (request.POST.get("all") == "1")
        qty = int(request.POST.get("qty") or 0)

        if not all_flag:
            if qty <= 0:
                messages.error(request, "Cantidad inválida.")
                return redirect("inventory:articulo_list")
            qs = qs[:min(qty, total)]

        ids = list(qs.values_list("articulo_id", flat=True))
        Articulo.objects.filter(articulo_id__in=ids).update(estado=Articulo.Estado.BAJA)

        costo = Decimal(getattr(Producto.objects.only("costo").get(pk=g["product_id"]), "costo", 0) or 0)
        articulos = list(Articulo.objects.select_related("product_id").filter(articulo_id__in=ids))
        baja = BajaStock.objects.create(usuario=request.user, local=local or "")
        movs = [
            MovimientoStock(
                tipo=MovimientoStock.Tipo.BAJA,
                local=local,
                usuario=request.user,
                articulo=a,
                producto=a.product_id,
                sku=a.sku,
                barcode=a.barcode,
                talle=a.talle,
                color=a.color,
                cantidad=1,
                costo_unitario=costo,
                baja=baja,
                precio_unitario=None,
                profit_unitario=Decimal("0.00"),
                nota="Baja por lote",
            )
            for a in articulos
        ]
        MovimientoStock.objects.bulk_create(movs)

        messages.success(request, f"Baja aplicada a {len(ids)} artículo(s).")
        return redirect("inventory:articulo_list")
    
from django.db.models import Q, Count, Max
from django.views.generic import ListView, View
from django.http import JsonResponse

class ArticuloListView(LoginRequiredMixin, ListView):
    model = Articulo
    template_name = "inventory/articulo/articulo_list.html"
    context_object_name = "articulos"
    paginate_by = 20

    def get_queryset(self):
        qs = _articulos_visibles_qs(self.request)

        estado = (self.request.GET.get("estado") or "DISP").strip().upper()
        if estado in ["DISP", "VEND", "BAJA"]:
            qs = qs.filter(estado=estado)

        scan = (self.request.GET.get("scan") or "").strip()
        if scan:
            return qs.filter(barcode=scan, estado="DISP").order_by("created_at", "articulo_id")

        q = (self.request.GET.get("q") or "").strip()
        field = (self.request.GET.get("field") or "all").strip().lower()

        allowed_fields = ["all", "sku", "barcode", "producto", "marca", "color", "talle", "id"]
        if field not in allowed_fields:
            field = "all"

        if (not self.request.user.is_staff) and (not q):
            return qs.none()

        if q:
            if field == "all":
                filt = (
                    Q(sku__icontains=q) |
                    Q(color__icontains=q) |
                    Q(barcode__icontains=q) |
                    Q(product_id__nombre__icontains=q) |
                    Q(product_id__marca__nombre__icontains=q)
                )
                if q.isdigit():
                    n = int(q)
                    filt |= Q(talle=n) | Q(articulo_id=n)
                qs = qs.filter(filt)

            elif field == "sku":
                qs = qs.filter(sku__icontains=q)
            elif field == "barcode":
                qs = qs.filter(barcode__icontains=q)
            elif field == "producto":
                qs = qs.filter(product_id__nombre__icontains=q)
            elif field == "marca":
                qs = qs.filter(product_id__marca__nombre__icontains=q)
            elif field == "color":
                qs = qs.filter(color__icontains=q)
            elif field == "talle":
                qs = qs.filter(talle=int(q)) if q.isdigit() else qs.none()
            elif field == "id":
                qs = qs.filter(articulo_id=int(q)) if q.isdigit() else qs.none()

        mode = (self.request.GET.get("mode") or "qty").strip().lower()
        if mode not in ["qty", "unit"]:
            mode = "qty"

        if mode == "unit":
            return qs.order_by("created_at", "articulo_id")

        values_fields = [
            "sku",
            "barcode",
            "talle",
            "color",
            "estado",
            "product_id",
            "product_id__nombre",
            "product_id__marca__nombre",
        ]

        if _should_show_all_locals(self.request):
            values_fields.extend([
                "local_id",
                "local_id__nombre",
            ])

        return (
            qs.values(*values_fields)
            .annotate(
                qty=Count("articulo_id"),
                last_created=Max("created_at"),
            )
            .order_by("-last_created", "barcode", "talle", "color")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["locales"] = Local.objects.all().order_by("nombre")
        ctx["all_locals_active"] = _should_show_all_locals(self.request)
        ctx["scan"] = (self.request.GET.get("scan") or "").strip()
        ctx["q"] = (self.request.GET.get("q") or "").strip()

        field = (self.request.GET.get("field") or "all").strip().lower()
        if field not in ["all", "sku", "producto", "marca", "color", "talle", "id", "barcode"]:
            field = "all"
        ctx["field"] = field

        ctx["estado"] = (self.request.GET.get("estado") or "DISP").upper()

        mode = (self.request.GET.get("mode") or "qty").strip().lower()
        if ctx["scan"]:
            mode = "unit"
        if mode not in ["qty", "unit"]:
            mode = "qty"
        ctx["mode"] = mode

        ctx["auto_open_first"] = bool(ctx["scan"] and ctx["mode"] == "unit" and ctx["articulos"])
        return ctx
    
def build_sku(producto, color: str, talle) -> str:
    nombre = (getattr(producto, "nombre", "") or "").strip()
    color = (color or "").strip()
    talle = str(talle).strip()
    sku = f"{nombre} {color} {talle}".strip()

    sku = " ".join(sku.split())  

    # truncar por seguridad (tu campo sku en DB suele ser 80)
    return sku[:80]

class ArticuloCreateView(LoginRequiredMixin, FormView):
    template_name = "inventory/articulo/articulo_create.html"
    form_class = ArticuloCreateForm
    success_url = reverse_lazy("inventory:articulo_list")
    
    def get_initial(self):
        
        initial = super().get_initial()
        local = _get_local_activo(self.request)
        barcode = (self.request.GET.get("barcode") or "").strip()

        if local and barcode:
            art = (
                Articulo.objects
                .select_related("product_id")
                .filter(local=local, barcode=barcode)
                .order_by("-articulo_id")
                .first()
            )
            if art:
                initial.update({
                    "barcode": barcode,
                    "product_id": art.product_id,
                    "talle": art.talle,
                    "color": art.color,
                    "sku_preview": build_sku(art.product_id, art.color, art.talle),
                })
        return initial
    
    @transaction.atomic
    def form_valid(self, form):
        producto = form.cleaned_data['product_id']
        barcode = form.cleaned_data['barcode']
        talle = form.cleaned_data['talle']
        color = (form.cleaned_data['color'] or "").strip()
        cantidad = form.cleaned_data['cantidad']
        local = _get_local_activo(self.request)
        referencia = (form.cleaned_data.get('referencia') or "").strip()
        costo_unitario = Decimal(getattr(producto, "costo", 0) or 0)
        sku = build_sku(producto, color, talle)

        existente = (
            Articulo.objects
            .select_for_update()
            .filter(local=local, barcode=barcode)
            .order_by("-articulo_id")
            .first()
        )

        if existente:
            if (
                existente.product_id_id != producto.pk
                or existente.talle != talle
                or (existente.color or "").strip().lower() != color.lower()
            ):
                form.add_error("barcode", "Ese barcode ya existe y pertenece a otro artículo.")
                return self.form_invalid(form)

        ingreso = Ingreso.objects.create(
            usuario=self.request.user,
            local=local,
            referencia=referencia,
            nota="Ingreso por carga de artículos"
        )

        total_linea = costo_unitario * cantidad
        item = IngresoItem.objects.create(
            ingreso=ingreso,
            producto=producto,
            sku=sku,
            barcode=barcode,
            talle=talle,
            color=color,
            cantidad=cantidad,
            costo_unitario=costo_unitario,
            total_linea=total_linea
        )

        articulos = [
            Articulo(
                product_id=producto,
                sku=sku,
                barcode=barcode,
                estado=Articulo.Estado.DISPONIBLE,
                talle=talle,
                color=color,
                local=local,
                ingreso_item=item
            )
            for _ in range(cantidad)
        ]
        Articulo.objects.bulk_create(articulos)

        creados = list(
            Articulo.objects
            .filter(ingreso_item=item, local=local)
            .order_by("articulo_id")[:cantidad]
        )

        movs = [
            MovimientoStock(
                tipo=MovimientoStock.Tipo.INGRESO,
                local=local,
                usuario=self.request.user,
                articulo=a,
                producto=producto,
                sku=sku,
                barcode=barcode,
                talle=talle,
                color=color,
                cantidad=1,
                costo_unitario=costo_unitario,
                precio_unitario=None,
                profit_unitario=Decimal("0.00"),
                ingreso=ingreso,
                venta=None,
                nota=f"Ingreso #{ingreso.ingreso_id}",
            )
            for a in creados
        ]
        MovimientoStock.objects.bulk_create(movs)

        messages.success(
            self.request,
            f"Ingreso #{ingreso.ingreso_id} registrado: {cantidad} unidad(es) de {producto}."
        )
        return super().form_valid(form)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["local"] = _get_local_activo(self.request)
        return kwargs


def _articulos_visibles_qs(request):
    qs = Articulo.objects.select_related("product_id", "product_id__marca", "local")

    if not _should_show_all_locals(request):
        local_id = request.session.get("local_id")
        if local_id:
            qs = qs.filter(local_id=local_id)

    return qs

class ArticuloLookupByBarcodeView(LoginRequiredMixin, View):
    def get(self, request):
        barcode = (request.GET.get("barcode") or "").strip()
        if not barcode:
            return JsonResponse({"found": False}, status=400)

        qs = _articulos_visibles_qs(request)

        art = (
            qs.filter(barcode=barcode)
              .order_by("-articulo_id")
              .first()
        )

        if not art:
            return JsonResponse({"found": False})

        return JsonResponse({
            "found": True,
            "product_id": art.product_id.product_id,
            "talle": art.talle,
            "color": art.color,
        })

def _norm(s):
    return (str(s).strip() if s is not None else "")

def _to_int(v, default=0):
    try:
        return int(v)
    except Exception:
        return default

class ArticuloImportXlsxView(FormView):
    template_name = "inventory/articulo/articulo_import_xlsx.html"
    form_class = ArticuloImportXlsxForm
    success_url = reverse_lazy("inventory:articulo_list")

    @transaction.atomic
    def form_valid(self, form):
        local = _get_local_activo(self.request)
        if not local:
            messages.error(self.request, "No hay un local seleccionado.")
            return redirect("inventory:articulo_list")

        f = form.cleaned_data["file"]

        try:
            wb = load_workbook(filename=f, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(self.request, "Error al leer documento Excel. Asegurate que sea .xlsx válido.")
            return self.form_invalid(form)

        # Leer encabezados
        header = []
        for cell in ws[1]:
            header.append(_norm(cell.value).lower())

        required = ["barcode", "producto_nombre", "talle", "color", "cantidad"]
        missing = [c for c in required if c not in header]
        if missing:
            messages.error(self.request, f"Faltan columnas: {', '.join(missing)}")
            return self.form_invalid(form)

        idx = {name: header.index(name) for name in header if name}

        creados_total = 0
        filas_ok = 0
        errores = []

        # Procesar filas
        ingreso = Ingreso.objects.create(
                usuario=self.request.user,
                local=local,
                referencia="IMPORT_XLSX",
                nota="Ingreso por importación Excel"
            )
        for row_num in range(2, ws.max_row + 1):
            row = [ws.cell(row=row_num, column=col).value for col in range(1, ws.max_column + 1)]

            barcode = _norm(row[idx["barcode"]])
            producto_nombre = _norm(row[idx["producto_nombre"]])
            talle = _to_int(row[idx["talle"]])
            color = _norm(row[idx["color"]])
            cantidad = _to_int(row[idx["cantidad"]], 0)
            referencia = _norm(row[idx["referencia"]]) if "referencia" in idx else ""

            if not barcode and not producto_nombre and not talle and not color and not cantidad:
                continue

            if not barcode:
                errores.append(f"Fila {row_num}: barcode vacío")
                continue
            if not producto_nombre:
                errores.append(f"Fila {row_num}: producto vacío")
                continue
            producto = Producto.objects.filter(nombre=producto_nombre).first()
            if not producto:
                errores.append(f"Fila {row_num}: no existe producto con nombre '{producto_nombre}'")
                continue
            if cantidad <= 0:
                errores.append(f"Fila {row_num}: cantidad inválida")
                continue
            
            conflicto = _barcode_conflict(local, barcode, producto, talle, color)
            if conflicto:
                errores.append(
                    f"Fila {row_num}: el barcode '{barcode}' ya existe y está asociado a "
                    f"'{conflicto.product_id}' / Color: '{conflicto.color}' / Talle: '{conflicto.talle}'"
                )
                continue
            
            # ---- MISMA LÓGICA QUE TU form_valid ----
            costo_unitario = Decimal(getattr(producto, "costo", 0) or 0)
            sku = build_sku(producto, color, talle)

            

            total_linea = costo_unitario * Decimal(cantidad)
            item = IngresoItem.objects.create(
                ingreso=ingreso,
                producto=producto,
                sku=sku,
                barcode=barcode,
                talle=talle,
                color=color,
                cantidad=cantidad,
                costo_unitario=costo_unitario,
                total_linea=total_linea
            )

            articulos = [
                Articulo(
                    product_id=producto,
                    sku=sku,
                    barcode=barcode,
                    estado=Articulo.Estado.DISPONIBLE,
                    talle=talle,
                    color=color,
                    local=local,
                    ingreso_item=item
                )
                for _ in range(cantidad)
            ]
            Articulo.objects.bulk_create(articulos)

            creados = list(
                Articulo.objects
                .filter(ingreso_item=item, local=local)
                .order_by("articulo_id")[:cantidad]
            )

            movs = [
                MovimientoStock(
                    tipo=MovimientoStock.Tipo.INGRESO,
                    local=local,
                    usuario=self.request.user,
                    articulo=a,
                    producto=producto,
                    sku=sku,
                    barcode=barcode,
                    talle=talle,
                    color=color,
                    cantidad=1,
                    costo_unitario=costo_unitario,
                    precio_unitario=None,
                    profit_unitario=Decimal("0.00"),
                    ingreso=ingreso,
                    venta=None,
                    nota=f"Ingreso #{ingreso.ingreso_id}",
                )
                for a in creados
            ]
            MovimientoStock.objects.bulk_create(movs)

            filas_ok += 1
            creados_total += cantidad

        # Feedback
        if filas_ok:
            messages.success(self.request, f"Importación OK: {filas_ok} fila(s), {creados_total} artículo(s).")
        if errores:
            # no lo hagas eterno; mostramos las primeras 10
            preview = "\n".join(errores[:10])
            messages.warning(self.request, f"Hubo errores en algunas filas:\n{preview}")

        return super().form_valid(form)

def _norm(value):
    return " ".join(str(value or "").strip().split())


def _to_decimal(value):
    if value is None or value == "":
        return None
    try:
        txt = str(value).strip().replace("$", "").replace(" ", "")
        txt = txt.replace(".", "").replace(",", ".") if "," in txt and "." in txt else txt.replace(",", ".")
        return Decimal(txt)
    except (InvalidOperation, ValueError):
        return None


class ProductoImportXlsxView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    template_name = "inventory/producto/producto_import_xlsx.html"
    form_class = ProductoImportXlsxForm
    success_url = reverse_lazy("inventory:producto_list")

    def test_func(self):
        return self.request.user.is_staff

    def form_valid(self, form):
        f = form.cleaned_data["file"]

        try:
            wb = load_workbook(filename=f, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(self.request, "No se pudo leer el Excel. Verificá que sea un .xlsx válido.")
            return self.form_invalid(form)

        header = [_norm(cell.value).lower() for cell in ws[1]]

        required = ["nombre", "tipo_producto", "material", "marca", "precio", "costo"]
        missing = [col for col in required if col not in header]
        if missing:
            messages.error(self.request, f"Faltan columnas obligatorias: {', '.join(missing)}")
            return self.form_invalid(form)

        idx = {name: header.index(name) for name in header if name}

        creados = 0
        actualizados = 0
        errores = []

        for row_num in range(2, ws.max_row + 1):
            row = [ws.cell(row=row_num, column=col).value for col in range(1, ws.max_column + 1)]

            nombre = _norm(row[idx["nombre"]])
            tipo_producto = _norm(row[idx["tipo_producto"]])
            material = _norm(row[idx["material"]])
            marca_nombre = _norm(row[idx["marca"]])
            precio = _to_decimal(row[idx["precio"]])
            costo = _to_decimal(row[idx["costo"]])

            if not any([nombre, tipo_producto, material, marca_nombre, precio, costo]):
                continue

            if not nombre:
                errores.append(f"Fila {row_num}: nombre vacío")
                continue

            if not tipo_producto:
                errores.append(f"Fila {row_num}: tipo_producto vacío")
                continue

            if not material:
                errores.append(f"Fila {row_num}: material vacío")
                continue

            if not marca_nombre:
                errores.append(f"Fila {row_num}: marca vacía")
                continue

            if precio is None:
                errores.append(f"Fila {row_num}: precio inválido")
                continue

            if costo is None:
                errores.append(f"Fila {row_num}: costo inválido")
                continue

            marca = Marca.objects.filter(nombre__iexact=marca_nombre).first()
            if not marca:
                errores.append(f"Fila {row_num}: no existe la marca '{marca_nombre}'")
                continue

            producto = Producto.objects.filter(nombre__iexact=nombre).first()

            if producto:
                producto.tipo_producto = tipo_producto
                producto.material = material
                producto.marca = marca
                producto.precio = precio
                producto.costo = costo
                producto.save()
                actualizados += 1
            else:
                Producto.objects.create(
                    nombre=nombre,
                    tipo_producto=tipo_producto,
                    material=material,
                    marca=marca,
                    precio=precio,
                    costo=costo,
                )
                creados += 1

        if creados or actualizados:
            messages.success(
                self.request,
                f"Importación completada. Creados: {creados}. Actualizados: {actualizados}."
            )

        if errores:
            preview = " | ".join(errores[:10])
            messages.warning(self.request, f"Se encontraron errores: {preview}")

        return super().form_valid(form)

def _norm_text(value):
    return " ".join(str(value or "").strip().split())

def _barcode_conflict(local, barcode, producto, talle, color):
    color_norm = _norm_text(color).lower()

    existente = (
        Articulo.objects
        .select_related("product_id")
        .filter(local=local, barcode=barcode)
        .order_by("-articulo_id")
        .first()
    )

    if not existente:
        return None

    mismo_producto = existente.product_id_id == producto.pk
    mismo_talle = existente.talle == talle
    mismo_color = _norm_text(existente.color).lower() == color_norm

    if mismo_producto and mismo_talle and mismo_color:
        return None

    return existente








CART_KEY = "pos_cart"

def _get_cart(session):
    return session.get(CART_KEY, {})

def _save_cart(session, cart):
    session[CART_KEY] = cart
    session.modified = True
    
def _cart_totals(cart):
    subtotal_base = Decimal("0.00")
    total_descuento = Decimal("0.00")
    total = Decimal("0.00")

    for it in cart.values():
        qty = int(it["qty"])
        precio_base = Decimal(it["precio_base"])
        descuento_total_linea = Decimal(it.get("descuento_total_linea", "0"))

        subtotal_linea = precio_base * qty
        total_linea = subtotal_linea - descuento_total_linea

        it["subtotal_bruto"] = str(subtotal_linea)
        it["total_linea"] = str(total_linea)

        subtotal_base += subtotal_linea
        total_descuento += descuento_total_linea
        total += total_linea

    return {
        "subtotal_base": subtotal_base,
        "total_descuento": total_descuento,
        "subtotal_final": total,
    }


def _get_local_activo(request):
    local_id = request.session.get("local_id")
    if not local_id:
        return None
    return Local.objects.get(local_id=local_id)
def _saldo_caja_local(local):
    if not local:
        return Decimal("0.00")

    ventas_efectivo = (
        Venta.objects
        .filter(
            local=local,
            estado=Venta.Estado.CERRADA,
            metodo_de_pago=Venta.MetodoPago.EFECTIVO,
        )
        .aggregate(total=Coalesce(Sum("total"), Decimal("0.00")))
    )["total"] or Decimal("0.00")

    entradas = (
        RetiroCaja.objects
        .filter(local=local, tipo=RetiroCaja.Tipo.ENTRADA)
        .aggregate(total=Coalesce(Sum("monto"), Decimal("0.00")))
    )["total"] or Decimal("0.00")

    salidas = (
        RetiroCaja.objects
        .filter(local=local, tipo=RetiroCaja.Tipo.SALIDA)
        .aggregate(total=Coalesce(Sum("monto"), Decimal("0.00")))
    )["total"] or Decimal("0.00")

    saldo = ventas_efectivo + entradas - salidas
    return saldo if saldo > 0 else Decimal("0.00")

class POSView(LoginRequiredMixin, View):
    template_name = "inventory/pos/pos.html"
    def get(self, request):
        cart = _get_cart(request.session)
        local=_get_local_activo(request)
        barcodes = [it["barcode"] for it in cart.values()]
        stock_map = {}
        if barcodes and local:
            rows = (Articulo.objects
                    .filter(local=local, barcode__in=barcodes, estado=Articulo.Estado.DISPONIBLE)
                    .values("barcode")
                    .annotate(disponibles=Count("articulo_id"))
                    )
            stock_map = {r["barcode"]: r["disponibles"] for r in rows}
        totals = _cart_totals(cart)
        hoy = timezone.localdate()  
        ventas_qs = (
            Venta.objects
            .filter(
                usuario=request.user,
                local=local if local else None,
                estado=Venta.Estado.CERRADA,
                fecha__date=hoy,   
            )
            .order_by("-venta_id")
        )
        agg = ventas_qs.aggregate(
            ventas_hoy_count=Count("venta_id"),
            ventas_hoy_total=Coalesce(Sum("total"), Decimal("0.00")),
        )
        ventas_hoy = ventas_qs.values("venta_id", "fecha", "metodo_de_pago", "total")[:50]
        total_en_efectivo = _saldo_caja_local(local)
            
        return render(request, self.template_name, {
            "cart": cart,
            "stock_map": stock_map,
            "subtotal_base": totals["subtotal_base"],
            "subtotal": totals["subtotal_final"],
            "total": totals["subtotal_final"],
            "total_descuento": totals["total_descuento"],
            "total_en_efectivo": total_en_efectivo,
            "local_activo": local,
            "hoy": hoy,
            "ventas_hoy": list(ventas_hoy),
            "ventas_hoy_count": agg["ventas_hoy_count"] or 0,
            "ventas_hoy_total": agg["ventas_hoy_total"] or Decimal("0.00"),
            
        })    

class POSCajaView(LoginRequiredMixin, View):
    template_name = "inventory/pos/pos_caja.html"
    def get(self, request):
        local = _get_local_activo(request)
        hoy = timezone.localdate()

        saldo_actual = _saldo_caja_local(local)

        movimientos_hoy = (
            RetiroCaja.objects
            .filter(local=local, fecha=hoy)
            .select_related("usuario")
            .order_by("-creado_en")[:30]
        )

        entradas_hoy = (
            RetiroCaja.objects
            .filter(local=local, fecha=hoy, tipo=RetiroCaja.Tipo.ENTRADA)
            .aggregate(total=Coalesce(Sum("monto"), Decimal("0.00")))
        )["total"] or Decimal("0.00")

        salidas_hoy = (
            RetiroCaja.objects
            .filter(local=local, fecha=hoy, tipo=RetiroCaja.Tipo.SALIDA)
            .aggregate(total=Coalesce(Sum("monto"), Decimal("0.00")))
        )["total"] or Decimal("0.00")

        return render(request, self.template_name, {
            "local_activo": local,
            "hoy": hoy,
            "saldo_actual": saldo_actual,
            "movimientos_hoy": movimientos_hoy,
            "entradas_hoy": entradas_hoy,
            "salidas_hoy": salidas_hoy,
            "puede_agregar_efectivo": request.user.is_staff,
        })
        
    def post(self, request):
        local = _get_local_activo(request)
        if not local:
            messages.error(request, "No hay un local activo seleccionado.")
            return redirect("inventory:pos_caja")

        hoy = timezone.localdate()

        accion = (request.POST.get("accion") or "salida").strip().lower()
        monto_raw = (request.POST.get("monto") or "").strip()
        nota = (request.POST.get("nota") or "").strip()
        motivo = (request.POST.get("motivo") or RetiroCaja.Motivo.OTRO).strip().upper()

        try:
            monto = Decimal(monto_raw)
        except (InvalidOperation, TypeError):
            messages.error(request, "Monto inválido.")
            return redirect("inventory:pos_caja")

        if monto <= 0:
            messages.error(request, "El monto debe ser mayor a 0.")
            return redirect("inventory:pos_caja")

        motivos_validos = {c for c, _ in RetiroCaja.Motivo.choices}
        if motivo not in motivos_validos:
            motivo = RetiroCaja.Motivo.OTRO

        if accion == "entrada":
            if not request.user.is_staff:
                messages.error(request, "Solo un administrador puede agregar efectivo.")
                return redirect("inventory:pos_caja")

            if motivo == RetiroCaja.Motivo.GUARDAR:
                motivo = RetiroCaja.Motivo.APORTE

            RetiroCaja.objects.create(
                local=local,
                usuario=request.user,
                fecha=hoy,
                tipo=RetiroCaja.Tipo.ENTRADA,
                monto=monto,
                motivo=motivo,
                nota=nota,
            )

            messages.success(request, f"Ingreso de efectivo registrado: $ {monto:.2f}")
            return redirect("inventory:pos_caja")

        saldo_disponible = _saldo_caja_local(local)
        if monto > saldo_disponible:
            messages.error(
                request,
                f"No alcanza el efectivo disponible. Disponible: $ {saldo_disponible:.2f}"
            )
            return redirect("inventory:pos_caja")

        RetiroCaja.objects.create(
            local=local,
            usuario=request.user,
            fecha=hoy,
            tipo=RetiroCaja.Tipo.SALIDA,
            monto=monto,
            motivo=motivo,
            nota=nota,
        )

        messages.success(request, f"Retiro registrado: $ {monto:.2f}")
        return redirect("inventory:pos_caja")

class POSAddItemByBarcodeView(LoginRequiredMixin, View):
    def post(self, request):
        barcode = (request.POST.get("barcode") or "").strip()
        if not barcode:
            messages.error(request, "Escaneá un código de barras.")
            return redirect("inventory:pos")

        local = _get_local_activo(request)
        if not local:
            messages.error(request, "No hay un local activo seleccionado.")
            return redirect("inventory:pos")

        qs = (
            Articulo.objects
            .select_related("product_id", "product_id__marca")
            .filter(
                local=local,
                barcode=barcode,
                estado=Articulo.Estado.DISPONIBLE
            )
            .order_by("articulo_id")
        )

        art = qs.first()
        if not art:
            messages.error(request, f"No hay unidades disponibles para el código {barcode}.")
            return redirect("inventory:pos")

        disponibles = qs.count()
        line_key = f"{barcode}|{art.talle}|{art.color}"
        cart = _get_cart(request.session)

        current_qty = int(cart.get(line_key, {}).get("qty", 0))
        new_qty = current_qty + 1

        if new_qty > disponibles:
            messages.error(request, f"No hay stock suficiente. Disponibles: {disponibles}.")
            return redirect("inventory:pos")

        precio_base = Decimal(art.product_id.precio)
        promo, promo_data = get_mejor_promocion_para_producto(art.product_id, qty=new_qty)
        descuento_total_linea = promo_data["descuento"] if promo_data else Decimal("0")
        precio_final_unit = promo_data["precio_final"] if promo_data else precio_base

        subtotal_bruto = precio_base * new_qty
        total_linea = subtotal_bruto - descuento_total_linea

        marca_obj = getattr(art.product_id, "marca", None)
        marca_nombre = getattr(marca_obj, "nombre", "") if marca_obj else ""

        cart[line_key] = {
            "producto_id": art.product_id.product_id,
            "producto_nombre": art.product_id.nombre,
            "marca": marca_nombre,
            "sku": art.sku,
            "barcode": art.barcode,
            "talle": art.talle,
            "color": art.color,

            "qty": new_qty,

            "precio_base": str(precio_base),
            "precio_final_unit": str(precio_final_unit),   # útil para mostrar promos normales
            "descuento_total_linea": str(descuento_total_linea),
            "subtotal_bruto": str(subtotal_bruto),
            "total_linea": str(total_linea),

            "promocion_id": promo.promocion_id if promo else None,
            "promocion_nombre": promo.nombre if promo else "",
        }

        _save_cart(request.session, cart)
        return redirect("inventory:pos")

class POSRemoveItemView(LoginRequiredMixin, View):
    def post(self, request):
        key = request.POST.get("key") or ""
        cart = _get_cart(request.session)
        if key in cart:
            del cart[key]
            _save_cart(request.session, cart)
        return redirect("inventory:pos")
    
class POSClearView(LoginRequiredMixin, View):
    def get(self, request):
        _save_cart(request.session, {})
        return redirect("inventory:pos")
    
    def post(self, request):
        _save_cart(request.session, {})
        return redirect("inventory:pos")
    
class POSCheckoutView(LoginRequiredMixin, View):
    template_name = "inventory/pos/pos_checkout.html"

    def get(self, request):
        form = CheckoutForm()
        cart = _get_cart(request.session)
        if not cart:
            messages.info(request, "El carrito está vacío.")
            return redirect("inventory:pos")

        totals = _cart_totals(cart)

        return render(request, self.template_name, {
            "form": form,
            "cart": cart,
            "subtotal_base": totals["subtotal_base"],
            "total_descuento": totals["total_descuento"],
            "total": totals["subtotal_final"],
        })

    @transaction.atomic
    def post(self, request):
        cart = _get_cart(request.session)
        if not cart:
            messages.error(request, "El carrito está vacío.")
            return redirect("inventory:pos")

        local = _get_local_activo(request)
        if not local:
            messages.error(request, "No hay un local activo seleccionado.")
            return redirect("inventory:pos")

        form = CheckoutForm(request.POST)
        if not form.is_valid():
            totals = _cart_totals(cart)
            return render(request, self.template_name, {
                "form": form,
                "cart": cart,
                "subtotal_base": totals["subtotal_base"],
                "total_descuento": totals["total_descuento"],
                "total": totals["subtotal_final"],
            })

        metodo_de_pago = form.cleaned_data["metodo_pago"]

        venta = Venta.objects.create(
            usuario=request.user,
            local=local,
            estado=Venta.Estado.ABIERTA,
            metodo_de_pago=metodo_de_pago
        )

        subtotal_base = Decimal("0.00")
        subtotal_final = Decimal("0.00")
        descuento_total = Decimal("0.00")
        profit_total = Decimal("0.00")

        for it in cart.values():
            barcode = it["barcode"]
            qty = int(it["qty"])
            qty_d = Decimal(qty)

            precio_base = Decimal(it["precio_base"])
            precio_final_unit = Decimal(it.get("precio_final_unit", it["precio_base"]))
            descuento_total_linea = Decimal(it.get("descuento_total_linea", "0"))
            promocion_id = it.get("promocion_id")
            promocion_nombre = it.get("promocion_nombre", "")

            disponibles_qs = (
                Articulo.objects
                .select_for_update()
                .filter(local=local, barcode=barcode, estado=Articulo.Estado.DISPONIBLE)
                .order_by("articulo_id")
            )

            articulos = list(disponibles_qs[:qty])
            if len(articulos) < qty:
                raise ValueError(
                    f"Stock insuficiente para barcode {barcode}. Pediste {qty}, hay {len(articulos)}."
                )

            costo_total = sum(
                Decimal(getattr(a.ingreso_item, "costo_unitario", 0) or 0)
                for a in articulos
            )
            costo_unitario_prom = (costo_total / qty_d) if qty else Decimal("0.00")

            total_linea_base = precio_base * qty_d
            total_linea = total_linea_base - descuento_total_linea
            precio_unitario_real = (total_linea / qty_d) if qty else Decimal("0.00")
            profit_linea = total_linea - costo_total

            subtotal_base += total_linea_base
            subtotal_final += total_linea
            descuento_total += descuento_total_linea
            profit_total += profit_linea

            promo_obj = None
            if promocion_id:
                promo_obj = Promocion.objects.filter(pk=promocion_id).first()

            descuento_unitario_prom = (descuento_total_linea / qty_d) if qty else Decimal("0.00")

            VentaItem.objects.create(
                venta=venta,
                producto_id=it["producto_id"],
                sku=it["sku"],
                barcode=barcode,
                talle=int(it["talle"]),
                color=it["color"],
                cantidad=qty,

                precio_base_unitario=precio_base,
                descuento_unitario=descuento_unitario_prom,
                precio_unitario=precio_unitario_real,

                costo_unitario=costo_unitario_prom,
                profit_linea=profit_linea,
                total_linea=total_linea,

                promocion=promo_obj,
                promocion_nombre=promocion_nombre,
            )

            movs = []
            venta_articulos = []

            for a in articulos:
                costo_u = Decimal(a.ingreso_item.costo_unitario) if a.ingreso_item else Decimal("0.00")
                profit_u = precio_unitario_real - costo_u

                movs.append(MovimientoStock(
                    tipo=MovimientoStock.Tipo.VENTA,
                    local=local,
                    usuario=request.user,
                    articulo=a,
                    producto=a.product_id,
                    barcode=a.barcode,
                    sku=a.sku,
                    talle=a.talle,
                    color=a.color,
                    cantidad=-1,
                    costo_unitario=costo_u,
                    precio_unitario=precio_unitario_real,
                    profit_unitario=profit_u,
                    ingreso=None,
                    venta=venta,
                    nota=f"Venta #{venta.venta_id}",
                ))

                venta_articulos.append(VentaArticulo(venta=venta, articulo=a))

            Articulo.objects.filter(
                articulo_id__in=[a.articulo_id for a in articulos]
            ).update(estado=Articulo.Estado.VENDIDO)

            VentaArticulo.objects.bulk_create(venta_articulos)
            MovimientoStock.objects.bulk_create(movs)

        venta.subtotal = subtotal_base
        venta.total_descuento = descuento_total
        venta.total = subtotal_final
        venta.profit_total = profit_total
        venta.estado = Venta.Estado.CERRADA
        venta.save(update_fields=[
            "subtotal",
            "total_descuento",
            "total",
            "profit_total",
            "estado",
        ])

        _save_cart(request.session, {})

        messages.success(request, f"Venta #{venta.venta_id} cerrada. Total: $ {venta.total}")
        return redirect("inventory:pos")  
    
class MovimientoStockView(LoginRequiredMixin, ListView):
    template_name = "inventory/movimientos/movimientos_list.html"

    def get(self, request):
        show_all = request.user.is_staff and (request.GET.get("all_locals") == "1")
        local = _get_local_activo(request)

        if not show_all and not local:
            return render(request, self.template_name, {
                "error_local": True,
                "mode": "doc",
                "rows": [],
                "tipo": "all",
                "q": "",
                "from": "",
                "to": "",
                "all_locals_active": False,
                "can_export_pdf": False,
            })

        mode = (request.GET.get("mode") or "doc").strip().lower()
        if mode not in ["doc", "day"]:
            mode = "doc"

        if mode == "day" and not request.user.is_staff:
            mode = "doc"

        tipo = (request.GET.get("tipo") or "all").strip().upper()
        q = (request.GET.get("q") or "").strip()
        desde = (request.GET.get("from") or "").strip()
        hasta = (request.GET.get("to") or "").strip()

        base = MovimientoStock.objects.select_related(
            "local", "usuario", "producto", "producto__marca", "venta", "ingreso", "articulo"
        )

        if not show_all:
            base = base.filter(local=local)

        if not request.user.is_staff:
            base = base.filter(usuario=request.user)

        if tipo in ["IN", "OUT", "TRF", "BAJ", "RET"]:
            base = base.filter(tipo=tipo)

        if q:
            base = base.filter(
                Q(barcode__icontains=q) |
                Q(sku__icontains=q) |
                Q(producto__nombre__icontains=q) |
                Q(producto__marca__nombre__icontains=q)
            )

        if desde:
            d = Datetime.strptime(desde, "%Y-%m-%d").date()
            base = base.filter(
                fecha__gte=timezone.make_aware(datetime.combine(d, time.min))
            )

        if hasta:
            h = Datetime.strptime(hasta, "%Y-%m-%d").date()
            base = base.filter(
                fecha__lte=timezone.make_aware(datetime.combine(h, time.max))
            )

        if mode == "doc":
            vals = ["tipo", "venta_id", "ingreso_id", "transferencia_id", "baja_id"]
            if show_all:
                vals += ["local_id", "local__nombre"]

            rows = (
                base.values(*vals)
                .annotate(
                    fecha=Max("fecha"),
                    items=Count("movimiento_id"),
                    unidades=Coalesce(Sum("cantidad"), Value(0, output_field=IntegerField())),
                    venta_total=Coalesce(Sum("precio_unitario"), Value(0, output_field=DecimalField())),
                    ganancia_total=Coalesce(Sum("profit_unitario"), Value(0, output_field=DecimalField())),
                )
                .order_by("-fecha")
            )

            return render(request, self.template_name, {
                "mode": mode,
                "rows": list(rows),
                "tipo": tipo,
                "q": q,
                "from": desde,
                "to": hasta,
                "all_locals_active": show_all,
                "can_export_pdf": request.user.is_staff and mode == "doc" and tipo in ["IN", "OUT"],
            })

        qs = base.annotate(dia=TruncDate("fecha"))
        vals = ["dia"]
        if show_all:
            vals += ["local_id", "local__nombre"]

        rows = (
            qs.values(*vals)
            .annotate(
                unidades_out=Coalesce(
                    Sum("cantidad", filter=Q(tipo="OUT")),
                    Value(0, output_field=IntegerField())
                ),
                unidades_in=Coalesce(
                    Sum("cantidad", filter=Q(tipo="IN")),
                    Value(0, output_field=IntegerField())
                ),
                venta_total=Coalesce(
                    Sum("precio_unitario", filter=Q(tipo="OUT")),
                    Value(0, output_field=DecimalField())
                ),
                ganancia_total=Coalesce(
                    Sum("profit_unitario", filter=Q(tipo="OUT")),
                    Value(0, output_field=DecimalField())
                ),
            )
            .order_by("-dia")
        )

        return render(request, self.template_name, {
            "mode": mode,
            "rows": list(rows),
            "tipo": tipo,
            "q": q,
            "from": desde,
            "to": hasta,
            "all_locals_active": show_all,
            "can_export_pdf": False,
        })

def money(x):
    if x is None:
        x = Decimal("0.00")
    return f"$ {Decimal(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def safe(s, maxlen=80):
    s = "" if s is None else str(s)
    return s[:maxlen]


def ensure_space(c, w, h, y, min_y=70, repeat_header_fn=None):
    if y < min_y:
        c.showPage()
        if repeat_header_fn:
            return repeat_header_fn()
        return h - 60
    return y

AR_TZ = ZoneInfo("America/Argentina/Buenos_Aires")

def ar_dt(dt):
    if dt is None:
        return None

    # Si Django lo tiene como aware, convertir a hora Argentina
    if timezone.is_aware(dt):
        return dt.astimezone(AR_TZ)

    # Si viene naive, asumir que ya está en hora local del negocio
    return dt

def fmt_ar_dt(dt):
    dt = ar_dt(dt)
    return dt.strftime("%Y-%m-%d %H:%M") if dt else "-"

def fmt_ar_date(dt):
    dt = ar_dt(dt)
    return dt.strftime("%Y-%m-%d") if dt else "-"

def _draw_centered_header(c, w, h, title, subtitle_lines):
    top = h - 42

    c.setFont("Helvetica-Bold", 15)
    c.drawCentredString(w / 2, top, title)

    y = top - 18
    c.setFont("Helvetica", 9)
    for line in subtitle_lines:
        c.drawCentredString(w / 2, y, line)
        y -= 12

    y -= 2
    c.setLineWidth(0.8)
    c.line(40, y, w - 40, y)

    return y - 18


def _draw_sales_summary_block(c, w, y, ventas_count, unidades, subtotal, descuentos, total, costo, resultado):
    c.setFont("Helvetica-Bold", 13)
    c.drawString(40, y, "Resumen de ventas")
    y -= 18

    c.setFont("Helvetica", 11)
    c.drawString(40, y, f"Comprobantes: {ventas_count}")
    y -= 14
    c.drawString(40, y, f"Unidades: {unidades}")
    y -= 16

    c.drawString(40, y, f"Subtotal: {money(subtotal)}")
    y -= 14
    c.drawString(40, y, f"Descuentos: {money(descuentos)}")
    y -= 14
    c.drawString(40, y, f"Costo total: {money(costo)}")
    y -= 18
    c.setFont("Helvetica-Bold", 11)
    c.drawString(40, y, f"Total vendido: {money(total)}")
    y -= 14
    c.drawString(40, y, f"Ganancia: {money(resultado)}")
    y -= 30

    return y


def _draw_sale_title(c, w, y, venta):
    c.setFont("Helvetica-Bold", 11)
    c.drawString(40, y, f"Venta {fmt_ar_dt(venta.fecha)}")
    y -= 14

    usuario_txt = ""
    if hasattr(venta.usuario, "get_full_name"):
        usuario_txt = venta.usuario.get_full_name().strip()
    if not usuario_txt:
        usuario_txt = venta.usuario.username

    c.setFont("Helvetica", 9)
    c.drawString(40, y, f"Usuario: {usuario_txt}   Método de pago: {venta.get_metodo_de_pago_display()}")
    y -= 10

    c.setLineWidth(0.6)
    c.line(40, y, w - 40, y)
    y -= 14

    return y


def _draw_sale_table_header(c, w, y):
    cols = [
        (40,  "SKU", "L"),
        (270, "Código", "L"),
        (355, "Cant.", "R"),
        (430, "Precio u.", "R"),
        (485, "Desc.", "R"),
        (w - 40, "Total", "R"),
    ]

    c.setFont("Helvetica-Bold", 9)
    for x, txt, align in cols:
        if align == "R":
            c.drawRightString(x, y, txt)
        else:
            c.drawString(x, y, txt)

    y -= 14
    c.setFont("Helvetica", 9)
    return y

def _draw_sale_totals(c, w, y, venta, costo_total, ganancia_total):
    c.setLineWidth(0.6)
    c.line(315, y, w - 40, y)
    y -= 18

    c.setFont("Helvetica-Bold", 9)
    c.drawRightString(
        w - 40,
        y,
        f"Subtotal: {money(venta.subtotal)}   Desc.: {money(venta.total_descuento)}   Costo: {money(costo_total)}"
    )
    y -= 20

    c.drawRightString(
        w - 40,
        y,
        f"Total: {money(venta.total)}   Ganancia: {money(ganancia_total)}"
    )
    y -= 30

    return y

def _estimate_sale_block_height(venta):
    items = list(venta.items.all())
    h = 34   # título + usuario + línea
    h += 16  # table header
    for it in items:
        h += 12
        descuento_u = Decimal(getattr(it, "descuento_unitario", 0) or 0)
        promo_obj = getattr(it, "promocion", None)
        promo_nombre = getattr(promo_obj, "nombre", "") if promo_obj else ""
        if descuento_u > 0 and promo_nombre:
            h += 10
    h += 6
    h += 34  # totales
    h += 18  # aire entre comprobantes
    return h

def _render_ventas_pdf(
    request,
    local,
    ventas_qs,
    include_caja=False,
    caja_fecha=None,
    caja_usuario=None,
):
    ventas = list(
        ventas_qs
        .select_related("usuario", "local")
        .prefetch_related("items", "items__producto", "items__producto__marca", "items__promocion")
        .order_by("-fecha", "-venta_id")
    )

    if not ventas:
        return HttpResponse("No hay ventas para exportar.", status=400)

    venta_ids = [v.venta_id for v in ventas]

    ventas_agg = (
        Venta.objects
        .filter(venta_id__in=venta_ids)
        .aggregate(
            ventas=Count("venta_id"),
            subtotal_bruto=Coalesce(Sum("subtotal"), Value(0, output_field=DecimalField())),
            descuento_total=Coalesce(Sum("total_descuento"), Value(0, output_field=DecimalField())),
            venta_total=Coalesce(Sum("total"), Value(0, output_field=DecimalField())),
        )
    )

    ms_global_agg = (
        MovimientoStock.objects
        .filter(venta_id__in=venta_ids, tipo=MovimientoStock.Tipo.VENTA)
        .aggregate(
            unidades=Count("movimiento_id"),
            costo_total=Coalesce(Sum("costo_unitario"), Value(0, output_field=DecimalField())),
            profit_total=Coalesce(Sum("profit_unitario"), Value(0, output_field=DecimalField())),
        )
    )

    fechas = [ar_dt(v.fecha) for v in ventas if v.fecha]
    min_fecha = min(fechas)
    max_fecha = max(fechas)

    if min_fecha.date() == max_fecha.date():
        fecha_line = f"Fecha: {fmt_ar_date(min_fecha)}"
    else:
        fecha_line = f"Rango: {fmt_ar_date(min_fecha)} a {fmt_ar_date(max_fecha)}"

    usuarios = sorted({
        (v.usuario.get_full_name().strip() if hasattr(v.usuario, "get_full_name") else "") or v.usuario.username
        for v in ventas
    })
    usuarios_txt = ", ".join(usuarios) if usuarios else "-"

    caja_rows = []
    caja_total = Decimal("0.00")

    if include_caja and caja_fecha:
        caja_qs = (
            RetiroCaja.objects
            .filter(local=local, fecha=caja_fecha)
            .select_related("usuario")
            .order_by("-creado_en")
        )

        if caja_usuario and not request.user.is_staff:
            caja_qs = caja_qs.filter(usuario=caja_usuario)

        caja_rows = list(caja_qs)

        caja_total = (
            caja_qs.aggregate(
                total=Coalesce(Sum("monto"), Value(0, output_field=DecimalField()))
            )["total"] or Decimal("0.00")
        )

    resp = HttpResponse(content_type="application/pdf")
    if len(ventas) == 1 and not include_caja:
        filename = f'venta_{ventas[0].venta_id}.pdf'
    elif include_caja:
        filename = "resumen_dia.pdf"
    else:
        filename = "reporte_ventas.pdf"
    resp["Content-Disposition"] = f'inline; filename="{filename}"'

    c = canvas.Canvas(resp, pagesize=A4)
    w, h = A4

    def header():
        title = "Resumen del Día" if include_caja else "Reporte de Ventas - FreeDonna"
        return _draw_centered_header(
            c, w, h,
            title=title,
            subtitle_lines=[
                f"Local: {local.nombre}",
                fecha_line,
                f"Usuarios: {safe(usuarios_txt, 110)}",
                f"Generado: {timezone.now().astimezone(AR_TZ).strftime('%Y-%m-%d %H:%M')}",
            ]
        )

    y = header()

    if len(ventas) > 1:
        y = ensure_space(c, w, h, y, min_y=150, repeat_header_fn=header)
        y = _draw_sales_summary_block(
            c, w, y,
            ventas_count=ventas_agg["ventas"] or 0,
            unidades=ms_global_agg["unidades"] or 0,
            subtotal=ventas_agg["subtotal_bruto"],
            descuentos=ventas_agg["descuento_total"],
            total=ventas_agg["venta_total"],
            costo=ms_global_agg["costo_total"],
            resultado=ms_global_agg["profit_total"],
        )

    for venta in ventas:
        estimated_height = _estimate_sale_block_height(venta)

        if y < max(120, 60 + estimated_height):
            c.showPage()
            y = header()

        venta_ms_agg = (
            MovimientoStock.objects
            .filter(venta=venta, tipo=MovimientoStock.Tipo.VENTA)
            .aggregate(
                costo_total=Coalesce(Sum("costo_unitario"), Value(0, output_field=DecimalField())),
                profit_total=Coalesce(Sum("profit_unitario"), Value(0, output_field=DecimalField())),
            )
        )

        y = _draw_sale_title(c, w, y, venta)
        y = _draw_sale_table_header(c, w, y)

        items = list(venta.items.all().order_by("item_id"))

        for it in items:
            if y < 95:
                c.showPage()
                y = header()
                y = _draw_sale_title(c, w, y, venta)
                y = _draw_sale_table_header(c, w, y)

            sku_base = safe(getattr(it, "sku", ""), 20)
            talle = safe(getattr(it, "talle", ""), 8)
            color = safe(getattr(it, "color", ""), 12)
            sku = safe(f"{sku_base} {color} {talle}".strip(), 34)

            barcode = safe(getattr(it, "barcode", ""), 24)
            qty = int(getattr(it, "cantidad", 1) or 1)

            precio_u = Decimal(getattr(it, "precio_unitario", 0) or 0)
            descuento_u = Decimal(getattr(it, "descuento_unitario", 0) or 0)
            total_linea = Decimal(getattr(it, "total_linea", 0) or 0)
            descuento_total = descuento_u * qty

            c.setFont("Helvetica", 9)
            c.drawString(40, y, sku)
            c.drawString(270, y, barcode)
            c.drawRightString(355, y, str(qty))
            c.drawRightString(430, y, money(precio_u))
            c.drawRightString(485, y, money(descuento_total))
            c.drawRightString(w - 40, y, money(total_linea))
            y -= 12

            promo_obj = getattr(it, "promocion", None)
            promo_nombre = getattr(promo_obj, "nombre", "") if promo_obj else ""
            if descuento_u > 0 and promo_nombre:
                c.setFont("Helvetica-Oblique", 8)
                c.drawString(40, y, safe(f"Promo aplicada: {promo_nombre}", 100))
                y -= 11

        y -= 6
        y = _draw_sale_totals(
            c, w, y,
            venta=venta,
            costo_total=venta_ms_agg["costo_total"],
            ganancia_total=venta_ms_agg["profit_total"],
        )

    if include_caja:
        y = ensure_space(c, w, h, y, min_y=140, repeat_header_fn=header)

        c.setStrokeColor(colors.HexColor("#D9DEE7"))
        c.line(40, y, w - 40, y)
        y -= 18

        c.setFont("Helvetica-Bold", 11)
        c.setFillColor(colors.HexColor("#1F2A44"))
        c.drawString(40, y, "Gastos de caja del día")
        y -= 16

        c.setFont("Helvetica", 9)
        c.setFillColor(colors.HexColor("#5B657A"))
        c.drawString(40, y, f"Cantidad de movimientos: {len(caja_rows)}")
        c.drawRightString(w - 40, y, f"Total gastos: {money(caja_total)}")
        y -= 18

        if caja_rows:
            c.setFont("Helvetica-Bold", 9)
            c.setFillColor(colors.HexColor("#1F2A44"))
            c.drawString(40, y, "Hora")
            c.drawString(95, y, "Usuario")
            c.drawRightString(w - 40, y, "Monto")
            y -= 10

            c.setStrokeColor(colors.HexColor("#D9DEE7"))
            c.line(40, y, w - 40, y)
            y -= 14

            for mov in caja_rows:
                y = ensure_space(c, w, h, y, min_y=90, repeat_header_fn=header)

                fecha_mov = getattr(mov, "creado_en", None) or getattr(mov, "fecha_hora", None)
                hora = ar_dt(fecha_mov).strftime("%H:%M") if fecha_mov else "--:--"

                usuario_txt = "-"
                if getattr(mov, "usuario", None):
                    usuario_txt = (
                        mov.usuario.get_full_name().strip()
                        if hasattr(mov.usuario, "get_full_name") and mov.usuario.get_full_name().strip()
                        else mov.usuario.username
                    )
                monto = getattr(mov, "monto", Decimal("0.00")) or Decimal("0.00")
                motivo = getattr(mov, "motivo", "") or ""
                nota = getattr(mov, "nota", "") or ""

                # fila principal
                c.setFont("Helvetica", 9)
                c.setFillColor(colors.black)
                c.drawString(40, y, hora)
                c.drawString(95, y, safe(usuario_txt, 28))
                c.drawString(230, y, safe(motivo, 25))   # 👈 motivo
                c.drawRightString(w - 40, y, money(monto))
                y -= 12

                
                if nota:
                    c.setFont("Helvetica-Oblique", 8)
                    c.setFillColor(colors.HexColor("#5B657A"))
                    c.drawString(95, y, safe(nota, 80))
                    y -= 11
        else:
            c.setFont("Helvetica", 9)
            c.setFillColor(colors.HexColor("#5B657A"))
            c.drawString(40, y, "No se registraron gastos de caja en el día.")
            y -= 14

    c.save()
    return resp

def venta_pdf(request, venta_id: int):
    local = _get_local_activo(request)

    venta = get_object_or_404(
        Venta.objects.select_related("local", "usuario"),
        venta_id=venta_id
    )

    if local and venta.local_id != local.local_id:
        return HttpResponse("No autorizado para este local.", status=403)

    ventas_qs = Venta.objects.filter(venta_id=venta.venta_id, local=venta.local)
    return _render_ventas_pdf(request, venta.local, ventas_qs)

def movimiento_pdf(request):
    if not request.user.is_staff:
        return HttpResponse("No autorizado.", status=403)

    local = _get_local_activo(request)
    if not local:
        return HttpResponse("No hay local activo.", status=400)

    mode = (request.GET.get("mode") or "doc").strip().lower()
    tipo = (request.GET.get("tipo") or "all").strip().upper()
    q = (request.GET.get("q") or "").strip()
    desde = (request.GET.get("from") or "").strip()
    hasta = (request.GET.get("to") or "").strip()

    base = (
        MovimientoStock.objects
        .filter(local=local)
        .select_related("producto", "venta", "ingreso", "articulo", "usuario")
    )

    if tipo in ["IN", "OUT", "ADJ", "TRF", "BAJ", "RET"]:
        base = base.filter(tipo=tipo)

    if q:
        base = base.filter(
            Q(barcode__icontains=q) |
            Q(sku__icontains=q) |
            Q(producto__nombre__icontains=q) |
            Q(producto__marca__nombre__icontains=q)
        )

    if desde:
        base = base.filter(fecha__date__gte=desde)
    if hasta:
        base = base.filter(fecha__date__lte=hasta)

    is_sales_report = (mode == "doc" and tipo in ["OUT", "VENTA"])
    if is_sales_report:
        venta_ids = list(
            base.exclude(venta_id=None)
                .values_list("venta_id", flat=True)
                .distinct()
        )

        ventas_qs = Venta.objects.filter(venta_id__in=venta_ids, local=local)
        return _render_ventas_pdf(request, local, ventas_qs)

    is_ingresos_report = (mode == "doc" and tipo in ["IN", "INGRESO"])
    if is_ingresos_report:
        ingreso_ids = list(
            base.exclude(ingreso_id=None)
                .values_list("ingreso_id", flat=True)
                .distinct()
        )

        ingresos_qs = Ingreso.objects.filter(ingreso_id__in=ingreso_ids, local=local)
        return _render_ingresos_pdf(request, local, ingresos_qs)

    if mode == "unit":
        rows = list(base.order_by("-fecha", "-movimiento_id")[:1500])
    elif mode == "day":
        rows = list(
            base.annotate(dia=TruncDate("fecha"))
                .values("dia")
                .annotate(
                    movimientos=Count("movimiento_id"),
                    unidades=Sum("cantidad"),
                    costo_total=Coalesce(Sum("costo_unitario"), Value(0, output_field=DecimalField())),
                    venta_total=Coalesce(Sum("precio_unitario"), Value(0, output_field=DecimalField())),
                    profit_total=Coalesce(Sum("profit_unitario"), Value(0, output_field=DecimalField())),
                )
                .order_by("-dia")
        )
    elif mode == "variant":
        rows = list(
            base.values("barcode", "sku", "talle", "color", "producto__nombre", "producto__marca__nombre")
                .annotate(
                    movimientos=Count("movimiento_id"),
                    unidades=Sum("cantidad"),
                    costo_total=Coalesce(Sum("costo_unitario"), Value(0, output_field=DecimalField())),
                    venta_total=Coalesce(Sum("precio_unitario"), Value(0, output_field=DecimalField())),
                    profit_total=Coalesce(Sum("profit_unitario"), Value(0, output_field=DecimalField())),
                    last_fecha=Max("fecha"),
                )
                .order_by("barcode", "talle", "color")
        )
    else:
        rows = list(
            base.values("tipo", "venta_id", "ingreso_id")
                .annotate(
                    fecha=Max("fecha"),
                    items=Count("movimiento_id"),
                    costo_total=Coalesce(Sum("costo_unitario"), Value(0, output_field=DecimalField())),
                    venta_total=Coalesce(Sum("precio_unitario"), Value(0, output_field=DecimalField())),
                    profit_total=Coalesce(Sum("profit_unitario"), Value(0, output_field=DecimalField())),
                )
                .order_by("-fecha")
        )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="movimientos.pdf"'
    c = canvas.Canvas(response, pagesize=A4)
    w, h = A4

    y = _draw_centered_header(
        c, w, h,
        title=f"Movimientos de Stock - {local.nombre}",
        subtitle_lines=[
            f"Modo: {mode}  |  Tipo: {tipo}",
            f"Rango: {desde or '-'} a {hasta or '-'}",
            f"Generado: {timezone.localtime().strftime('%Y-%m-%d %H:%M')}",
        ]
    )

    c.setFont("Helvetica", 9)
    c.drawString(40, y, "Render de movimientos general pendiente / conservar lógica actual.")
    c.save()
    return response

def pos_resumen_dia_pdf(request):
    local = _get_local_activo(request)
    if not local:
        return HttpResponse("No hay local activo.", status=400)

    hoy = timezone.localdate()

    ventas_qs = (
        Venta.objects
        .filter(
            usuario=request.user,
            local=local,
            estado=Venta.Estado.CERRADA,
            fecha__date=hoy,
        )
        .order_by("-fecha", "-venta_id")
    )

    return _render_ventas_pdf(
        request,
        local,
        ventas_qs,
        include_caja=True,
        caja_fecha=hoy,
        caja_usuario=request.user,
    )
    
def _get_resumen_dia_base(request):
    local = _get_local_activo(request)
    if not local:
        return None, None, None

    hoy = timezone.localdate()

    ventas_qs = (
        Venta.objects
        .select_related("usuario", "local")
        .filter(
            usuario=request.user,
            local=local,
            estado=Venta.Estado.CERRADA,
            fecha__date=hoy,
        )
        .order_by("-fecha", "-venta_id")
    )

    return local, hoy, ventas_qs


def _build_pos_resumen_dia_context(request):
    local, hoy, ventas_qs = _get_resumen_dia_base(request)
    if not local:
        return None

    ventas = list(ventas_qs)
    efectivo=_saldo_caja_local(local)
    ventas_agg = ventas_qs.aggregate(
        cantidad_ventas=Coalesce(Count("venta_id"), 0),
        total_vendido=Coalesce(Sum("total"), Value(0, output_field=DecimalField())),
    )

    movimientos_caja_qs = (
        RetiroCaja.objects
        .select_related("usuario", "local")
        .filter(
            usuario=request.user,
            local=local,
            fecha=hoy,
        )
        .order_by("-fecha", "-retiro_id")
    )

    movimientos_caja = list(movimientos_caja_qs)

    total_mov_caja = movimientos_caja_qs.aggregate(
        total=Coalesce(Sum("monto"), Value(0, output_field=DecimalField()))
    )["total"]

    resumen = {
        "cantidad_ventas": ventas_agg["cantidad_ventas"] or 0,
        "total_vendido": ventas_agg["total_vendido"] or 0,
        "caja_efectivo": efectivo,
        "neto_caja": total_mov_caja or 0,
    }
    email = "joaquindores@gmail.com"
    return {
        "local": local,
        "hoy": hoy,
        "ventas": ventas,
        "movimientos_caja": movimientos_caja,
        "resumen": resumen,
        "email_destino": email,
    }


def _generar_pdf_resumen_dia_response(request):
    local, hoy, ventas_qs = _get_resumen_dia_base(request)
    if not local:
        return HttpResponse("No hay local activo.", status=400)

    return _render_ventas_pdf(
        request,
        local,
        ventas_qs,
        include_caja=True,
        caja_fecha=hoy,
        caja_usuario=request.user,
    )


@login_required
def pos_resumen_dia_view(request):
    ctx = _build_pos_resumen_dia_context(request)
    if ctx is None:
        messages.error(request, "No hay local activo.")
        return redirect("inventory:pos")

    return render(request, "inventory/pos/export_caja.html", ctx)


@login_required
def pos_resumen_dia_enviar_view(request):
    if request.method != "POST":
        return redirect("inventory:pos_resumen_dia")

    ctx = _build_pos_resumen_dia_context(request)
    if ctx is None:
        messages.error(request, "No hay local activo.")
        return redirect("inventory:pos")

    local = ctx["local"]
    email_destino = ctx["email_destino"]

    if not email_destino:
        messages.error(request, "Este local no tiene configurado un email de destino.")
        return redirect("inventory:pos_resumen_dia")

    try:
        pdf_response = _generar_pdf_resumen_dia_response(request)
        if pdf_response.status_code != 200:
            messages.error(request, "No se pudo generar el PDF del resumen.")
            return redirect("inventory:pos_resumen_dia")

        pdf_bytes = pdf_response.content

        fecha_txt = timezone.localtime().strftime("%d-%m-%Y")
        nombre_local = getattr(local, "nombre", "local")

        email = EmailMessage(
            subject=f"Resumen del día - {nombre_local} - {fecha_txt}",
            body=(
                f"Se adjunta el resumen del día.\n\n"
                f"Local: {nombre_local}\n"
                f"Empleado: {request.user.get_username()}\n"
                f"Fecha: {fecha_txt}\n"
            ),
            to=[email_destino],
        )

        email.attach(
            f"resumen_dia_{nombre_local}_{fecha_txt}.pdf",
            pdf_bytes,
            "application/pdf"
        )
        email.send(fail_silently=False)

        messages.success(request, f"Resumen enviado a {email_destino}.")
    except Exception as e:
        messages.error(request, f"No se pudo enviar el resumen: {e}")

    return redirect("inventory:pos_resumen_dia")

class VentaDetailView(LoginRequiredMixin, DetailView):
    model = Venta
    template_name = "inventory/movimientos/venta_detail.html"
    context_object_name = "venta"
    pk_url_kwarg = "venta_id"
    
    def get_object(self, queryset=None):
        local = _get_local_activo(self.request)
        qs = Venta.objects.select_related("usuario", "local")
        obj = get_object_or_404(qs, venta_id=self.kwargs["venta_id"])
        if local and obj.local_id != local.local_id:
            raise get_object_or_404(Venta, venta_id=-1)
        return obj

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        venta = ctx["venta"]

        items = list(
            venta.items
            .select_related("producto", "producto__marca", "promocion")
            .all()
            .order_by("item_id")
        )
        ctx["items"] = items

        unidades = (
            MovimientoStock.objects
            .select_related("articulo", "producto")
            .filter(venta=venta, tipo=MovimientoStock.Tipo.VENTA)
            .order_by("movimiento_id")
        )
        ctx["unidades"] = unidades

        agg = (
            MovimientoStock.objects
            .filter(venta=venta, tipo=MovimientoStock.Tipo.VENTA)
            .aggregate(
                unidades=Count("movimiento_id"),
                costo_total=Coalesce(Sum("costo_unitario"), Decimal("0.00")),
                venta_total=Coalesce(Sum("precio_unitario"), Decimal("0.00")),
                profit_total=Coalesce(Sum("profit_unitario"), Decimal("0.00")),
            )
        )

        ctx["ms_totals"] = {
            "unidades": agg["unidades"] or 0,
            "costo_total": agg["costo_total"] or Decimal("0.00"),
            "venta_total": agg["venta_total"] or Decimal("0.00"),
            "profit_total": agg["profit_total"] or Decimal("0.00"),
        }

        ctx["sale_totals"] = {
            "subtotal": venta.subtotal or Decimal("0.00"),
            "descuento_total": getattr(venta, "total_descuento", Decimal("0.00")) or Decimal("0.00"),
            "total": venta.total or Decimal("0.00"),
            "profit_total": venta.profit_total or Decimal("0.00"),
        }

        return ctx
    
class IngresoDetailView(LoginRequiredMixin, DetailView):
    model = Ingreso
    template_name = "inventory/movimientos/ingreso_detail.html"
    context_object_name = "ingreso"
    pk_url_kwarg = "ingreso_id"

    def get_object(self, queryset=None):
        local = _get_local_activo(self.request)
        qs = Ingreso.objects.select_related("usuario", "local")
        obj = get_object_or_404(qs, ingreso_id=self.kwargs["ingreso_id"])
        if local and obj.local_id != local.local_id:
            raise get_object_or_404(Ingreso, ingreso_id=-1)
        return obj

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ingreso = ctx["ingreso"]

        items = list(ingreso.items.all().order_by("item_id"))
        ctx["items"] = items

        unidades = (MovimientoStock.objects
                    .select_related("articulo")
                    .filter(ingreso=ingreso, tipo=MovimientoStock.Tipo.INGRESO)
                    .order_by("movimiento_id"))
        ctx["unidades"] = unidades

        agg = (MovimientoStock.objects
               .filter(ingreso=ingreso, tipo=MovimientoStock.Tipo.INGRESO)
               .aggregate(
                   unidades=Count("movimiento_id"),
                   costo_total=Sum("costo_unitario"),
               ))
        ctx["ms_totals"] = {
            "unidades": agg["unidades"] or 0,
            "costo_total": agg["costo_total"] or Decimal("0.00"),
        }
        return ctx

class BajaDetailView(LoginRequiredMixin, DetailView):
    model = BajaStock
    template_name = "inventory/movimientos/baja_detail.html"
    context_object_name = "baja"
    pk_url_kwarg = "baja_id"

    def get_object(self, queryset=None):
        local = _get_local_activo(self.request)
        qs = BajaStock.objects.select_related("usuario", "local")
        obj = get_object_or_404(qs, baja_id=self.kwargs["baja_id"])
        if local and obj.local_id != local.local_id:
            raise get_object_or_404(BajaStock, baja_id=-1)
        return obj

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        baja = ctx["baja"]

        unidades = (MovimientoStock.objects
                    .select_related("articulo")
                    .filter(baja=baja, tipo=MovimientoStock.Tipo.BAJA)
                    .order_by("movimiento_id"))
        ctx["unidades"] = unidades

        items_map = {}
        total_unidades = 0
        total_costo = Decimal("0.00")

        for m in unidades:
            key = (m.sku, m.barcode, m.talle, m.color)
            qty = int(m.cantidad or 0)
            cu = Decimal(m.costo_unitario or 0)
            line = cu * qty

            total_unidades += qty
            total_costo += line

            if key not in items_map:
                items_map[key] = {
                    "sku": m.sku,
                    "barcode": m.barcode,
                    "talle": m.talle,
                    "color": m.color,
                    "cantidad": 0,
                    "total_linea": Decimal("0.00"),
                }

            items_map[key]["cantidad"] += qty
            items_map[key]["total_linea"] += line

        items = list(items_map.values())
        for it in items:
            if it["cantidad"] > 0:
                it["costo_unitario"] = (it["total_linea"] / it["cantidad"]).quantize(Decimal("0.01"))
            else:
                it["costo_unitario"] = Decimal("0.00")

        ctx["items"] = items
        ctx["ms_totals"] = {"unidades": total_unidades, "costo_total": total_costo}
        return ctx
    
def _draw_ingresos_header(c, w, h, title, subtitle_lines):
    top = h - 42

    c.setFont("Helvetica-Bold", 15)
    c.drawCentredString(w / 2, top, title)

    y = top - 18
    c.setFont("Helvetica", 9)
    for line in subtitle_lines:
        c.drawCentredString(w / 2, y, line)
        y -= 12

    y -= 2
    c.setLineWidth(0.8)
    c.line(40, y, w - 40, y)

    return y - 22
def _draw_ingresos_summary_block(c, y, ingresos_count, unidades, costo_total):
    c.setFont("Helvetica-Bold", 11)
    c.drawString(40, y, "Resumen del reporte")
    y -= 18

    c.setFont("Helvetica", 9)
    c.drawString(40, y, f"Comprobantes: {ingresos_count}")
    y -= 12
    c.drawString(40, y, f"Unidades: {unidades}")
    y -= 12
    c.drawString(40, y, f"Costo total: {money(costo_total)}")
    y -= 28

    return y

def _draw_ingreso_title(c, w, y, ingreso):
    c.setFont("Helvetica-Bold", 11)
    c.drawString(40, y, f"Ingreso {fmt_ar_dt(ingreso.fecha)}")
    y -= 14

    usuario_txt = ""
    if hasattr(ingreso.usuario, "get_full_name"):
        usuario_txt = ingreso.usuario.get_full_name().strip()
    if not usuario_txt:
        usuario_txt = ingreso.usuario.username

    c.setFont("Helvetica", 9)
    c.drawString(40, y, f"Usuario: {usuario_txt}")
    y -= 12

    if ingreso.referencia:
        c.drawString(40, y, f"Referencia: {safe(ingreso.referencia, 90)}")
        y -= 12

    if ingreso.nota:
        c.drawString(40, y, f"Nota: {safe(ingreso.nota, 95)}")
        y -= 12

    c.setLineWidth(0.6)
    c.line(40, y, w - 40, y)
    y -= 14

    return y

def _draw_ingreso_table_header(c, w, y):
    cols = [
        (40,  "SKU", "L"),
        (280, "Código", "L"),
        (430, "Cant.", "R"),
        (505, "Costo u.", "R"),
        (w - 40, "Total", "R"),
    ]

    c.setFont("Helvetica-Bold", 9)
    for x, txt, align in cols:
        if align == "R":
            c.drawRightString(x, y, txt)
        else:
            c.drawString(x, y, txt)

    y -= 14
    c.setFont("Helvetica", 9)
    return y

def _draw_ingreso_totals(c, w, y, ingreso, costo_total, unidades):
    c.setLineWidth(0.6)
    c.line(360, y, w - 40, y)
    y -= 16

    c.setFont("Helvetica-Bold", 9)
    c.drawRightString(w - 40, y, f"Costo total: {money(costo_total)}")
    y -= 14

    c.drawRightString(w - 40, y, f"Unidades: {unidades}")
    y -= 28

    return y

def _draw_ingreso_totals(c, w, y, ingreso, costo_total, unidades):
    c.setLineWidth(0.6)
    c.line(360, y, w - 40, y)
    y -= 16

    c.setFont("Helvetica-Bold", 9)
    c.drawRightString(w - 40, y, f"Costo total: {money(costo_total)}")
    y -= 14

    c.drawRightString(w - 40, y, f"Unidades: {unidades}")
    y -= 28

    return y

def _estimate_ingreso_block_height(ingreso):
    items = list(ingreso.items.all())
    h = 52  # título + usuario + línea base

    if ingreso.referencia:
        h += 12
    if ingreso.nota:
        h += 12

    h += 16  # header tabla
    h += 12 * len(items)
    h += 42  # cierre
    h += 18  # aire entre comprobantes

    return h

def _render_ingresos_pdf(request, local, ingresos_qs):
    ingresos = list(
        ingresos_qs
        .select_related("usuario", "local")
        .prefetch_related("items")
        .order_by("-fecha", "-ingreso_id")
    )

    if not ingresos:
        return HttpResponse("No hay ingresos para exportar.", status=400)

    ingreso_ids = [i.ingreso_id for i in ingresos]

    ingresos_agg = (
        Ingreso.objects
        .filter(ingreso_id__in=ingreso_ids)
        .aggregate(
            ingresos=Count("ingreso_id"),
        )
    )

    ms_global_agg = (
        MovimientoStock.objects
        .filter(ingreso_id__in=ingreso_ids, tipo=MovimientoStock.Tipo.INGRESO)
        .aggregate(
            unidades=Coalesce(Sum("cantidad"), Value(0)),
            costo_total=Coalesce(Sum("costo_unitario"), Value(0, output_field=DecimalField())),
        )
    )

    fechas = [ar_dt(i.fecha) for i in ingresos if i.fecha]
    min_fecha = min(fechas)
    max_fecha = max(fechas)

    if min_fecha.date() == max_fecha.date():
        fecha_line = f"Fecha: {fmt_ar_date(min_fecha)}"
    else:
        fecha_line = f"Rango: {fmt_ar_date(min_fecha)} a {fmt_ar_date(max_fecha)}"

    usuarios = sorted({
        (i.usuario.get_full_name().strip() if hasattr(i.usuario, "get_full_name") else "") or i.usuario.username
        for i in ingresos
    })
    usuarios_txt = ", ".join(usuarios) if usuarios else "-"

    resp = HttpResponse(content_type="application/pdf")
    if len(ingresos) == 1:
        filename = f'ingreso_{ingresos[0].ingreso_id}.pdf'
    else:
        filename = "reporte_ingresos.pdf"
    resp["Content-Disposition"] = f'inline; filename="{filename}"'

    c = canvas.Canvas(resp, pagesize=A4)
    w, h = A4

    def header():
        return _draw_ingresos_header(
            c, w, h,
            title="Reporte de Ingresos - FreeDonna",
            subtitle_lines=[
                f"Local: {local.nombre}",
                fecha_line,
                f"Usuarios: {safe(usuarios_txt, 110)}",
                f"Generado: {timezone.now().astimezone(AR_TZ).strftime('%Y-%m-%d %H:%M')}",
            ]
        )

    y = header()

    if len(ingresos) > 1:
        y = ensure_space(c, w, h, y, min_y=140, repeat_header_fn=header)
        y = _draw_ingresos_summary_block(
            c, y,
            ingresos_count=ingresos_agg["ingresos"] or 0,
            unidades=ms_global_agg["unidades"] or 0,
            costo_total=ms_global_agg["costo_total"] or 0,
        )

    for ingreso in ingresos:
        items = list(ingreso.items.all().order_by("item_id"))

        ingreso_ms_agg = (
            MovimientoStock.objects
            .filter(ingreso=ingreso, tipo=MovimientoStock.Tipo.INGRESO)
            .aggregate(
                unidades=Coalesce(Sum("cantidad"), Value(0)),
                costo_total=Coalesce(Sum("costo_unitario"), Value(0, output_field=DecimalField())),
            )
        )

        estimated_height = _estimate_ingreso_block_height(ingreso)
        if y - estimated_height < 70:
            c.showPage()
            y = header()

        y = _draw_ingreso_title(c, w, y, ingreso)
        y = _draw_ingreso_table_header(c, w, y)

        for it in items:
            if y < 95:
                c.showPage()
                y = header()
                y = _draw_ingreso_title(c, w, y, ingreso)
                y = _draw_ingreso_table_header(c, w, y)

            sku_base = safe(getattr(it, "sku", ""), 22)
            talle = safe(getattr(it, "talle", ""), 8)
            color = safe(getattr(it, "color", ""), 12)
            sku = safe(f"{sku_base} {color} {talle}".strip(), 38)

            barcode = safe(getattr(it, "barcode", ""), 22)
            qty = int(getattr(it, "cantidad", 1) or 1)

            costo_u = Decimal(getattr(it, "costo_unitario", 0) or 0)
            total_linea = Decimal(getattr(it, "total_linea", 0) or 0)

            c.setFont("Helvetica", 9)
            c.drawString(40, y, sku)
            c.drawString(280, y, barcode)
            c.drawRightString(430, y, str(qty))
            c.drawRightString(505, y, money(costo_u))
            c.drawRightString(w - 40, y, money(total_linea))
            y -= 12

        y -= 8
        y = _draw_ingreso_totals(
            c, w, y,
            ingreso=ingreso,
            costo_total=ingreso_ms_agg["costo_total"] or 0,
            unidades=ingreso_ms_agg["unidades"] or 0,
        )

    c.save()
    return resp

def ingreso_pdf(request, ingreso_id: int):
    local = _get_local_activo(request)

    ingreso = get_object_or_404(
        Ingreso.objects.select_related("local", "usuario"),
        ingreso_id=ingreso_id
    )

    if local and ingreso.local_id != local.local_id:
        return HttpResponse("No autorizado para este local.", status=403)

    ingresos_qs = Ingreso.objects.filter(ingreso_id=ingreso.ingreso_id, local=ingreso.local)
    return _render_ingresos_pdf(request, ingreso.local, ingresos_qs)

class ArticulosTransferirView(LoginRequiredMixin, View):
    
    @transaction.atomic
    def post(self, request):
        local_origen = _get_local_activo(request)
        if not local_origen:
            messages.error(request, "No hay un local activo seleccionado.")
            return redirect("inventory:articulo_list")

        destino_id = (request.POST.get("destino_id") or "").strip()
        barcode = (request.POST.get("barcode") or "").strip()
        nota_user = (request.POST.get("nota") or "").strip()
        qty_raw = (request.POST.get("qty") or "1").strip()

        if not destino_id or not barcode:
            messages.error(request, "Faltan datos para transferir (destino/barcode).")
            return redirect("inventory:articulo_list")

        try:
            qty = int(qty_raw)
        except ValueError:
            qty = 0

        if qty <= 0:
            messages.error(request, "La cantidad debe ser mayor a 0.")
            return redirect("inventory:articulo_list")

        destino = get_object_or_404(Local, local_id=destino_id)
        if destino.local_id == local_origen.local_id:
            messages.error(request, "El destino debe ser distinto al local origen.")
            return redirect("inventory:articulo_list")

        # Seleccionamos N artículos DISPONIBLES de ese barcode en el local origen
        qs = (Articulo.objects
              .select_for_update()
              .select_related("product_id")
              .filter(local=local_origen, barcode=barcode, estado=Articulo.Estado.DISPONIBLE)
              .order_by("articulo_id"))

        articulos = list(qs[:qty])
        if len(articulos) < qty:
            messages.error(request, f"Stock insuficiente. Pediste {qty} y hay {len(articulos)} disponibles.")
            return redirect("inventory:articulo_list")

        # Documento transferencia
        trf = Transferencia.objects.create(
            local_origen=local_origen,
            local_destino=destino,
            usuario=request.user,
            nota=nota_user,
        )

        # Items
        TransferenciaItem.objects.bulk_create([
            TransferenciaItem(
                transferencia=trf,
                articulo=a,
                sku=a.sku,
                barcode=a.barcode,
                talle=a.talle,
                color=a.color,
            ) for a in articulos
        ])

        # Mover artículos al destino (en batch)
        art_ids = [a.articulo_id for a in articulos]
        Articulo.objects.filter(articulo_id__in=art_ids).update(local=destino)

        # Movimientos: salida (origen) y entrada (destino)
        movs = []
        for a in articulos:
            nota = f"Transferencia #{trf.transferencia_id}: {local_origen.nombre} → {destino.nombre}. {nota_user}".strip()

            # salida
            movs.append(MovimientoStock(
                tipo=MovimientoStock.Tipo.TRANSFERENCIA,
                local=local_origen,
                local_origen=local_origen,
                local_destino=destino,
                transferencia=trf,
                usuario=request.user,
                articulo=a,
                producto=a.product_id,
                sku=a.sku,
                barcode=a.barcode,
                talle=a.talle,
                color=a.color,
                cantidad=-1,
                costo_unitario=Decimal("0.00"),
                precio_unitario=None,
                profit_unitario=Decimal("0.00"),
                ingreso=None,
                venta=None,
                nota=nota,
            ))
            # entrada
            movs.append(MovimientoStock(
                tipo=MovimientoStock.Tipo.TRANSFERENCIA,
                local=destino,
                local_origen=local_origen,
                local_destino=destino,
                transferencia=trf,
                usuario=request.user,
                articulo=a,
                producto=a.product_id,
                sku=a.sku,
                barcode=a.barcode,
                talle=a.talle,
                color=a.color,
                cantidad=+1,
                costo_unitario=Decimal("0.00"),
                precio_unitario=None,
                profit_unitario=Decimal("0.00"),
                ingreso=None,
                venta=None,
                nota=nota,
            ))

        MovimientoStock.objects.bulk_create(movs)

        messages.success(request, f"Transferencia #{trf.transferencia_id} realizada: {qty} unidad(es).")
        return redirect("inventory:transferencia_detail", transferencia_id=trf.transferencia_id)

class TransferenciaDetailView(LoginRequiredMixin, DetailView):
    model = Transferencia
    template_name = "inventory/movimientos/transferencia_detail.html"
    context_object_name = "trf"
    pk_url_kwarg = "transferencia_id"

    def get_object(self, queryset=None):
        local = _get_local_activo(self.request)
        qs = Transferencia.objects.select_related("local_origen", "local_destino", "usuario")
        obj = get_object_or_404(qs, transferencia_id=self.kwargs["transferencia_id"])
        # Solo ver si el local activo participa
        if local and (obj.local_origen_id != local.local_id and obj.local_destino_id != local.local_id):
            raise get_object_or_404(Transferencia, transferencia_id=-1)
        return obj

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        trf = ctx["trf"]
        ctx["items"] = list(trf.items.select_related("articulo").order_by("item_id"))
        return ctx


def _parse_pct(val: str):
    val = (val or "").strip()
    if val == "":
        return None
    try:
        return Decimal(val.replace(",", "."))
    except Exception:
        return "ERR"


def _apply_pct(value: Decimal, pct: Decimal) -> Decimal:
    factor = Decimal("1") + (pct / Decimal("100"))
    newv = (value * factor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if newv < 0:
        newv = Decimal("0.00")
    return newv

def _add_query_param(url, key, value):
    u = urlparse(url)
    q = dict(parse_qsl(u.query))
    q[key] = str(value)
    new_query = urlencode(q)
    return urlunparse((u.scheme, u.netloc, u.path, u.params, new_query, u.fragment))

class ProductoBulkAdjustPreviewView(LoginRequiredMixin, StaffRequiredMixin, View):
    def get(self, request):
        marca_id = (request.GET.get("marca_id") or "").strip()
        pct_precio = _parse_pct(request.GET.get("pct_precio"))
        pct_costo  = _parse_pct(request.GET.get("pct_costo"))

        if not marca_id:
            return JsonResponse({"ok": False, "error": "Seleccioná una marca."}, status=400)

        if pct_precio == "ERR" or pct_costo == "ERR":
            return JsonResponse({"ok": False, "error": "Porcentaje inválido."}, status=400)

        if pct_precio is None and pct_costo is None:
            return JsonResponse({"ok": False, "error": "Ingresá % en precio y/o costo."}, status=400)

        marca = get_object_or_404(Marca, pk=marca_id)
        qs = Producto.objects.filter(marca=marca).order_by("product_id")

        total = qs.count()
        sample = list(qs[:5])

        items = []
        for p in sample:
            old_precio = Decimal(p.precio or 0)
            old_costo  = Decimal(p.costo or 0)

            new_precio = _apply_pct(old_precio, pct_precio) if pct_precio is not None else old_precio
            new_costo  = _apply_pct(old_costo,  pct_costo)  if pct_costo  is not None else old_costo

            items.append({
                "id": p.pk,
                "nombre": getattr(p, "nombre", "") or str(p),
                "old_precio": f"{old_precio:.2f}",
                "new_precio": f"{new_precio:.2f}",
                "old_costo": f"{old_costo:.2f}",
                "new_costo": f"{new_costo:.2f}",
            })

        return JsonResponse({
            "ok": True,
            "marca": {"id": marca.pk, "nombre": marca.nombre},
            "total": total,
            "pct_precio": str(pct_precio) if pct_precio is not None else None,
            "pct_costo": str(pct_costo) if pct_costo is not None else None,
            "sample": items,
        })


class ProductoBulkAdjustApplyView(LoginRequiredMixin, StaffRequiredMixin, View):
    @transaction.atomic
    def post(self, request):
        marca_id = (request.POST.get("marca_id") or "").strip()
        pct_precio = _parse_pct(request.POST.get("pct_precio"))
        pct_costo  = _parse_pct(request.POST.get("pct_costo"))

        if not marca_id:
            messages.error(request, "Seleccioná una marca.")
            return redirect(request.META.get("HTTP_REFERER", "inventory:producto_list"))

        if pct_precio == "ERR" or pct_costo == "ERR":
            messages.error(request, "Porcentaje inválido.")
            return redirect(request.META.get("HTTP_REFERER", "inventory:producto_list"))

        if pct_precio is None and pct_costo is None:
            messages.error(request, "Ingresá % en precio y/o costo.")
            return redirect(request.META.get("HTTP_REFERER", "inventory:producto_list"))

        marca = get_object_or_404(Marca, pk=marca_id)
        qs = Producto.objects.filter(marca=marca).select_for_update()

        total = qs.count()
        if total == 0:
            messages.warning(request, f"No hay productos para {marca.nombre}.")
            return redirect(request.META.get("HTTP_REFERER", "inventory:producto_list"))

        adjust = ProductoBulkAdjust.objects.create(
            user=request.user,
            marca=marca,
            pct_precio=pct_precio,
            pct_costo=pct_costo,
            afectados=total,
            note="Ajuste masivo por marca desde pantalla de productos",
        )

        items_to_create = []
        # aplicamos y guardamos snapshot exacto
        for p in qs:
            old_precio = Decimal(p.precio or 0)
            old_costo  = Decimal(p.costo or 0)

            new_precio = _apply_pct(old_precio, pct_precio) if pct_precio is not None else old_precio
            new_costo  = _apply_pct(old_costo,  pct_costo)  if pct_costo  is not None else old_costo

            # Update producto
            p.precio = new_precio
            p.costo = new_costo
            p.save(update_fields=["precio", "costo"])

            items_to_create.append(ProductoBulkAdjustItem(
                adjust=adjust,
                producto=p,
                old_precio=old_precio,
                old_costo=old_costo,
                new_precio=new_precio,
                new_costo=new_costo,
            ))

        ProductoBulkAdjustItem.objects.bulk_create(items_to_create, batch_size=1000)

        # mensaje con link para deshacer
        undo_url = f"/inventario/productos/ajuste-marca/undo/{adjust.pk}/"  # o reverse() si preferís
        messages.success(
            request,
            f"Ajuste aplicado a {marca.nombre} ({total} productos). "
            f"Si necesitás revertirlo, abrí “Ajuste por marca” y tocá “Deshacer último ajuste”."
        )
        referer = request.META.get("HTTP_REFERER")
        fallback = redirect("inventory:producto_list").url  # o reverse(...)
        target = referer or fallback

        target = _add_query_param(target, "last_adjust", adjust.pk)
        target = _add_query_param(target, "last_brand", marca.nombre)
        return redirect(target)


class ProductoBulkAdjustUndoView(LoginRequiredMixin, StaffRequiredMixin, View):
    @transaction.atomic
    def post(self, request, adjust_id: int):
        adjust = get_object_or_404(ProductoBulkAdjust, pk=adjust_id)

        if adjust.estado == ProductoBulkAdjust.Estado.DESHECHO:
            messages.warning(request, "Este ajuste ya fue deshecho.")
            return redirect(request.META.get("HTTP_REFERER", "inventory:producto_list"))

        # bloqueamos items/productos
        items = list(adjust.items.select_related("producto").select_for_update())

        for it in items:
            p = it.producto
            # revertimos EXACTO a snapshot
            if it.old_precio is not None:
                p.precio = it.old_precio
            if it.old_costo is not None:
                p.costo = it.old_costo
            p.save(update_fields=["precio", "costo"])

        adjust.estado = ProductoBulkAdjust.Estado.DESHECHO
        adjust.save(update_fields=["estado"])

        messages.success(request, f"Ajuste deshecho. Los precios y costos volvieron al estado anterior.")
        return redirect(request.META.get("HTTP_REFERER", "inventory:producto_list"))

    # opcional: permitir GET con confirmación simple (yo prefiero POST)
    def get(self, request, adjust_id: int):
        return HttpResponseForbidden("Usá POST para deshacer.")

#promociones

def _round_money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def promocion_aplica_a_producto(promocion, producto):
    if not promocion.esta_vigente():
        return False

    if promocion.aplica_a_todos:
        return True

    if getattr(producto, "marca_id", None) and promocion.marcas.filter(pk=producto.marca_id).exists():
        return True

    if promocion.productos.filter(product_id=producto.product_id).exists():
        return True

    return False

def calcular_precio_con_promocion(producto, promo, qty=1):
    precio = Decimal(producto.precio)

    if promo.tipo_descuento == Promocion.TipoDescuento.PORCENTAJE:
        valor = promo.valor or Decimal("0")
        desc_unit = (valor / Decimal("100")) * precio
        return {
            "precio_final": precio - desc_unit,
            "descuento": desc_unit * qty,
        }

    if promo.tipo_descuento == Promocion.TipoDescuento.MONTO_FIJO:
        valor = promo.valor or Decimal("0")
        desc_unit = min(precio, valor)
        return {
            "precio_final": precio - desc_unit,
            "descuento": desc_unit * qty,
        }

    if promo.tipo_descuento == Promocion.TipoDescuento.ESCALON:
        unidad_obj = promo.unidad_objetivo or 0
        porc = promo.descuento_porcentaje or Decimal("0")

        if unidad_obj < 2 or porc <= 0:
            return {
                "precio_final": precio,
                "descuento": Decimal("0"),
            }

        if qty < unidad_obj:
            return {
                "precio_final": precio,
                "descuento": Decimal("0"),
            }

        descuento_una_unidad = (porc / Decimal("100")) * precio

        return {
            "precio_final": precio,
            "descuento": descuento_una_unidad,
        }

    return {
        "precio_final": precio,
        "descuento": Decimal("0"),
    }


def get_promociones_activas():
    ahora = timezone.now()
    qs = Promocion.objects.filter(estado=Promocion.Estado.ACTIVA).prefetch_related("marcas", "productos")

    # filtrado fino de fechas en Python por simplicidad y claridad
    return [p for p in qs if p.esta_vigente()]

def get_mejor_promocion_para_producto(producto, qty=1):
    promos = get_promociones_activas()

    mejor = None
    mejor_resultado = None

    for promo in promos:
        if not promocion_aplica_a_producto(promo, producto):
            continue

        resultado = calcular_precio_con_promocion(producto, promo, qty)

        if resultado["descuento"] <= Decimal("0"):
            continue

        if mejor is None:
            mejor = promo
            mejor_resultado = resultado
            continue

        if resultado["descuento"] > mejor_resultado["descuento"]:
            mejor = promo
            mejor_resultado = resultado
        elif resultado["descuento"] == mejor_resultado["descuento"]:
            if promo.prioridad > mejor.prioridad:
                mejor = promo
                mejor_resultado = resultado

    return mejor, mejor_resultado


class PromocionListView(LoginRequiredMixin, StaffRequiredMixin, ListView):
    model = Promocion
    template_name = "inventory/promocion/promocion_list.html"
    context_object_name = "promociones"
    paginate_by = 20

    def get_queryset(self):
        qs = Promocion.objects.prefetch_related("marcas", "productos").all()

        q = (self.request.GET.get("q") or "").strip()
        estado = (self.request.GET.get("estado") or "").strip().upper()

        if q:
            qs = qs.filter(
                Q(nombre__icontains=q) |
                Q(descripcion__icontains=q)
            )

        if estado in [Promocion.Estado.ACTIVA, Promocion.Estado.PAUSADA]:
            qs = qs.filter(estado=estado)

        return qs.order_by("-prioridad", "-created_at", "nombre")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = (self.request.GET.get("q") or "").strip()
        ctx["estado"] = (self.request.GET.get("estado") or "").strip().upper()
        return ctx


class PromocionCreateView(LoginRequiredMixin, StaffRequiredMixin, CreateView):
    model = Promocion
    form_class = PromocionForm
    template_name = "inventory/promocion/promocion_form.html"
    success_url = reverse_lazy("inventory:promocion_list")

    def form_valid(self, form):
        messages.success(self.request, "Promoción creada correctamente.")
        return super().form_valid(form)

    def form_invalid(self, form):
        print("FORM ERRORS:", form.errors)
        print("NON FIELD ERRORS:", form.non_field_errors())
        messages.error(self.request, "No se pudo guardar la promoción. Revisá los campos.")
        return super().form_invalid(form)

class PromocionUpdateView(LoginRequiredMixin, StaffRequiredMixin, UpdateView):
    model = Promocion
    form_class = PromocionForm
    template_name = "inventory/promocion/promocion_form.html"
    pk_url_kwarg = "promocion_id"
    success_url = reverse_lazy("inventory:promocion_list")

    def form_valid(self, form):
        messages.success(self.request, "Promoción actualizada correctamente.")
        return super().form_valid(form)


class PromocionDetailView(LoginRequiredMixin, StaffRequiredMixin, DetailView):
    model = Promocion
    template_name = "inventory/promocion/promocion_detail.html"
    context_object_name = "promocion"
    pk_url_kwarg = "promocion_id"


class PromocionToggleEstadoView(LoginRequiredMixin, StaffRequiredMixin, View):
    def post(self, request, promocion_id):
        promocion = get_object_or_404(Promocion, pk=promocion_id)

        if promocion.estado == Promocion.Estado.ACTIVA:
            promocion.estado = Promocion.Estado.PAUSADA
            msg = "Promoción pausada correctamente."
        else:
            promocion.estado = Promocion.Estado.ACTIVA
            msg = "Promoción activada correctamente."

        promocion.save(update_fields=["estado"])
        messages.success(request, msg)
        return redirect("inventory:promocion_list")
class PromocionDeleteView(LoginRequiredMixin, StaffRequiredMixin, View):
    def post(self, request, promocion_id):
        promo = get_object_or_404(Promocion, promocion_id=promocion_id)
        nombre = promo.nombre
        promo.delete()
        messages.success(request, f'Promoción "{nombre}" eliminada correctamente.')
        return redirect("inventory:promocion_list")