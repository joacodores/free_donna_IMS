from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse
from zoneinfo import ZoneInfo
from django.http import Http404, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, render, HttpResponse
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, View, DeleteView, TemplateView, FormView
from django.db.models import Q, Count, ExpressionWrapper, Sum, Max, Value, CharField, F, Case, When
from django.db.models.fields import DecimalField, IntegerField
from django.shortcuts import redirect
from httpcore import request
from sqlalchemy import Cast
from ..models import BajaStock, Ingreso, IngresoItem, Local, Marca, MovimientoStock, Producto, Articulo, ProductoBulkAdjust, ProductoBulkAdjustItem, Promocion, RetiroCaja, Transferencia, TransferenciaItem, Venta, VentaItem, VentaArticulo
from ..forms import ArticuloEditForm, ArticuloImportXlsxForm, CheckoutForm, ProductoImportXlsxForm, PromocionForm, TransferirArticuloForm, UserLoginForm, UserRegisterForm, ArticuloCreateForm, ArticuloImportXlsxForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic.edit import FormView
from django.db import transaction
from django.contrib import messages
from datetime import datetime as Datetime, time, timedelta, timezone, datetime
from django.db.models.functions import TruncDate, Coalesce, TruncMinute, Concat
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from django.utils import timezone
from openpyxl import load_workbook
from django.core.mail import EmailMessage

from .utilidades import (
    StaffRequiredMixin, _norm, _to_decimal, _norm_text, _should_show_all_locals,
    _get_local_activo
)


class ProductoListView(LoginRequiredMixin, ListView):
    model = Producto
    template_name = "inventory/producto/producto_list.html"
    context_object_name = "productos"
    paginate_by = 20  
    
    def get_queryset(self):
        qs = super().get_queryset().order_by("product_id")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(nombre__icontains=q)
        return qs
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = (self.request.GET.get("q") or "").strip()
        ctx["marcas"] = Marca.objects.all().order_by("nombre")
        return ctx  

class ProductoDetailView(DetailView):
    model = Producto
    template_name = "inventory/producto_list.html"
    context_object_name = "producto"
    
class ProductoCreateView(LoginRequiredMixin, StaffRequiredMixin,CreateView):
    model = Producto
    fields = ["nombre", "tipo_producto", "material", "marca", "precio", "costo"]
    template_name = "inventory/producto/producto_form.html"
    success_url = reverse_lazy("inventory:producto_list")


class ProductoUpdateView(LoginRequiredMixin, StaffRequiredMixin, UpdateView):
    model = Producto
    fields = ["nombre", "tipo_producto", "material", "marca", "precio", "costo"]
    template_name = "inventory/producto/producto_form.html"
    success_url = reverse_lazy("inventory:producto_list")


class ProductoDeleteView(LoginRequiredMixin,StaffRequiredMixin, View):
    model = Producto
    template_name = "inventory/producto/producto_confirm_delete.html"
    success_url = reverse_lazy("inventory:producto_list")
    
    def post(self, request, pk):
        producto = Producto.objects.get(pk=pk)
        producto.delete()
        return redirect(self.success_url)


class SetLocalView(LoginRequiredMixin, View):
    def post(self, request):
        local_id = request.POST.get("local_id")
        if Local.objects.filter(local_id=local_id).exists():
            request.session["local_id"] = int(local_id)
        return redirect(request.META.get("HTTP_REFERER", "/"))


class ProductoImportXlsxView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    template_name = "inventory/producto/producto_import_xlsx.html"
    form_class = ProductoImportXlsxForm
    success_url = reverse_lazy("inventory:producto_list")

    def test_func(self):
        return self.request.user.is_staff

    def form_valid(self, form):
        f = form.cleaned_data["file"]

        try:
            wb = load_workbook(filename=f, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(self.request, "No se pudo leer el Excel. Verificá que sea un .xlsx válido.")
            return self.form_invalid(form)

        header = [_norm(cell.value).lower() for cell in ws[1]]

        required = ["nombre", "tipo_producto", "material", "marca", "precio"]
        missing = [col for col in required if col not in header]
        if missing:
            messages.error(self.request, f"Faltan columnas obligatorias: {', '.join(missing)}")
            return self.form_invalid(form)

        idx = {name: header.index(name) for name in header if name}

        creados = 0
        actualizados = 0
        errores = []

        for row_num in range(2, ws.max_row + 1):
            row = [ws.cell(row=row_num, column=col).value for col in range(1, ws.max_column + 1)]

            nombre = _norm(row[idx["nombre"]])
            tipo_producto = _norm(row[idx["tipo_producto"]])
            material = _norm(row[idx["material"]])
            marca_nombre = _norm(row[idx["marca"]])
            precio = _to_decimal(row[idx["precio"]])
            costo_raw = row[idx["costo"]] if "costo" in idx else None
            costo_txt = _norm(costo_raw)
            costo = _to_decimal(costo_raw) if "costo" in idx else None

            if not any([nombre, tipo_producto, material, marca_nombre, precio, costo]):
                continue

            if not nombre:
                errores.append(f"Fila {row_num}: nombre vacío")
                continue

            if not tipo_producto:
                errores.append(f"Fila {row_num}: tipo_producto vacío")
                continue

            if not material:
                errores.append(f"Fila {row_num}: material vacío")
                continue

            if not marca_nombre:
                errores.append(f"Fila {row_num}: marca vacía")
                continue

            if precio is None:
                errores.append(f"Fila {row_num}: precio inválido")
                continue

            if costo_txt and costo is None:
                errores.append(f"Fila {row_num}: costo inválido")
                continue

            marca = Marca.objects.filter(nombre__iexact=marca_nombre).first()
            if not marca:
                errores.append(f"Fila {row_num}: no existe la marca '{marca_nombre}'")
                continue

            producto = Producto.objects.filter(nombre__iexact=nombre).first()

            if producto:
                producto.tipo_producto = tipo_producto
                producto.material = material
                producto.marca = marca
                producto.precio = precio
                producto.costo = costo
                producto.save()
                actualizados += 1
            else:
                Producto.objects.create(
                    nombre=nombre,
                    tipo_producto=tipo_producto,
                    material=material,
                    marca=marca,
                    precio=precio,
                    costo=costo,
                )
                creados += 1

        if creados or actualizados:
            messages.success(
                self.request,
                f"Importación completada. Creados: {creados}. Actualizados: {actualizados}."
            )

        if errores:
            preview = " | ".join(errores[:10])
            messages.warning(self.request, f"Se encontraron errores: {preview}")

        return super().form_valid(form)
