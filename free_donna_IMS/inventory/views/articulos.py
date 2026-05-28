from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse
from zoneinfo import ZoneInfo
from django.http import Http404, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, render, HttpResponse
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, View, DeleteView, TemplateView, FormView
from django.db.models import Q, Count, ExpressionWrapper, Sum, Max, Value, CharField, F, Case, When
from django.db.models.fields import DecimalField, IntegerField
from django.shortcuts import redirect
from httpcore import request
from sqlalchemy import Cast
from ..models import BajaStock, Ingreso, IngresoItem, Local, Marca, MovimientoStock, Producto, Articulo, ProductoBulkAdjust, ProductoBulkAdjustItem, Promocion, RetiroCaja, Transferencia, TransferenciaItem, Venta, VentaItem, VentaArticulo
from ..forms import ArticuloEditForm, ArticuloImportXlsxForm, CheckoutForm, ProductoImportXlsxForm, PromocionForm, TransferirArticuloForm, UserLoginForm, UserRegisterForm, ArticuloCreateForm, ArticuloImportXlsxForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic.edit import FormView
from django.db import transaction
from django.contrib import messages
from datetime import datetime as Datetime, time, timedelta, timezone, datetime
from django.db.models.functions import TruncDate, Coalesce, TruncMinute, Concat
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from django.utils import timezone
from openpyxl import load_workbook
from django.core.mail import EmailMessage

from .utilidades import (
    StaffRequiredMixin, _norm, _to_int, _barcode_conflict, _articulos_visibles_qs,
    _should_show_all_locals, _get_local_activo, build_sku
)


class ArticuloUpdateView(LoginRequiredMixin, StaffRequiredMixin, UpdateView):
    model = Articulo
    form_class = ArticuloEditForm
    context_object_name = "articulo"
    pk_url_kwarg = "articulo_id"
    template_name = "inventory/articulo/articulo_form.html"

    def get_success_url(self):
        return reverse_lazy("inventory:articulo_list")

class ArticulosBulkEditView(LoginRequiredMixin, StaffRequiredMixin, FormView):
    template_name = "inventory/articulo/articulo_bulk_form.html"
    form_class = ArticuloEditForm

    def dispatch(self, request, *args, **kwargs):
        local = _get_local_activo(request)
        if not local:
            messages.error(request, "No hay un local activo seleccionado.")
            return redirect("inventory:articulo_list")
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw["bulk"] = True
        return kw

    def _group(self):
        local = _get_local_activo(self.request)

        g = {
            "product_id": (self.request.GET.get("product_id") or self.request.POST.get("g_product_id") or "").strip(),
            "barcode": (self.request.GET.get("barcode") or self.request.POST.get("g_barcode") or "").strip(),
            "talle": (self.request.GET.get("talle") or self.request.POST.get("g_talle") or "").strip(),
            "color": (self.request.GET.get("color") or self.request.POST.get("g_color") or "").strip(),
            "sku": (self.request.GET.get("sku") or self.request.POST.get("g_sku") or "").strip(),
            "estado": (self.request.GET.get("estado") or self.request.POST.get("g_estado") or "DISP").strip().upper(),
        }

        if not all([g["product_id"], g["barcode"], g["talle"], g["color"], g["sku"], g["estado"]]):
            return None, None

        qs = Articulo.objects.filter(
            local=local,
            estado=g["estado"],
            product_id_id=g["product_id"],
            barcode=g["barcode"],
            talle=g["talle"],
            color=g["color"],
            sku=g["sku"],
        ).order_by("-articulo_id")

        return g, qs

    def get(self, request, *args, **kwargs):
        g, qs = self._group()
        if g is None:
            messages.error(request, "Faltan datos para editar por lote.")
            return redirect("inventory:articulo_list")
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        g, qs = self._group()
        ctx["group"] = g
        ctx["max_qty"] = qs.count() if qs is not None else 0
        return ctx

    @transaction.atomic
    def form_valid(self, form):
        g, qs = self._group()
        if g is None or qs is None:
            messages.error(self.request, "No se pudo resolver el grupo.")
            return redirect("inventory:articulo_list")

        if g["estado"] != "DISP":
            messages.error(self.request, "Solo podés editar por lote artículos DISP.")
            return redirect("inventory:articulo_list")

        all_flag = (self.request.POST.get("all") == "1")
        qty = int(self.request.POST.get("qty") or 0)
        total = qs.count()

        if total == 0:
            messages.warning(self.request, "No hay artículos para editar en ese grupo.")
            return redirect("inventory:articulo_list")

        if not all_flag:
            if qty <= 0:
                messages.error(self.request, "Cantidad inválida.")
                return redirect("inventory:articulo_list")
            qs = qs[:min(qty, total)]

        updates = {}
        for k in ["barcode", "product_id", "sku", "talle", "color"]:
            v = form.cleaned_data.get(k)
            if v not in (None, "", []):
                updates[k] = v

        ids = list(qs.values_list("articulo_id", flat=True))
        Articulo.objects.filter(articulo_id__in=ids).update(**updates)

        messages.success(self.request, f"Editados {len(ids)} artículo(s).")
        return redirect("inventory:articulo_list")
    
class ArticuloBajaView(LoginRequiredMixin, StaffRequiredMixin, View):
    @transaction.atomic
    def post(self, request, articulo_id):
        local = _get_local_activo(request)
        if not local:
            messages.error(request, "No hay un local activo seleccionado.")
            return redirect("inventory:articulo_list")

        a = get_object_or_404(Articulo, articulo_id=articulo_id, local=local)

        if a.estado != Articulo.Estado.DISPONIBLE:
            messages.error(request, "Solo podés dar de baja artículos DISP.")
            return redirect("inventory:articulo_list")

        a.estado = Articulo.Estado.BAJA
        a.save(update_fields=["estado"])
        baja = BajaStock.objects.create(usuario=request.user, local=local or "")
        MovimientoStock.objects.create(
            tipo=MovimientoStock.Tipo.BAJA,
            local=local,
            usuario=request.user,
            articulo=a,
            producto=a.product_id,
            sku=a.sku,
            barcode=a.barcode,
            talle=a.talle,
            color=a.color,
            cantidad=1,
            costo_unitario=getattr(a.product_id, "costo", None),
            precio_unitario=None,
            profit_unitario=Decimal("0.00"),
            baja=baja,
            nota="Baja manual",
        )

        messages.success(request, "Artículo dado de baja.")
        return redirect("inventory:articulo_list") 

class ArticulosBulkBajaView(LoginRequiredMixin, StaffRequiredMixin, View):
    @transaction.atomic
    def post(self, request):
        local = _get_local_activo(request)
        if not local:
            messages.error(request, "No hay un local activo seleccionado.")
            return redirect("inventory:articulo_list")

        g = {
            "product_id": (request.POST.get("g_product_id") or request.POST.get("product_id") or "").strip(),
            "barcode": (request.POST.get("g_barcode") or request.POST.get("barcode") or "").strip(),
            "talle": (request.POST.get("g_talle") or request.POST.get("talle") or "").strip(),
            "color": (request.POST.get("g_color") or request.POST.get("color") or "").strip(),
            "sku": (request.POST.get("g_sku") or request.POST.get("sku") or "").strip(),
            "estado": (request.POST.get("g_estado") or request.POST.get("estado") or "DISP").strip().upper(),
        }

        qs = Articulo.objects.filter(
            local=local,
            estado=Articulo.Estado.DISPONIBLE,
            product_id_id=g["product_id"],
            barcode=g["barcode"],
            talle=g["talle"],
            color=g["color"],
            sku=g["sku"],
        ).order_by("-articulo_id")

        total = qs.count()
        if total == 0:
            messages.warning(request, "No hay artículos para dar de baja.")
            return redirect("inventory:articulo_list")

        all_flag = (request.POST.get("all") == "1")
        qty = int(request.POST.get("qty") or 0)

        if not all_flag:
            if qty <= 0:
                messages.error(request, "Cantidad inválida.")
                return redirect("inventory:articulo_list")
            qs = qs[:min(qty, total)]

        ids = list(qs.values_list("articulo_id", flat=True))
        Articulo.objects.filter(articulo_id__in=ids).update(estado=Articulo.Estado.BAJA)

        costo = getattr(Producto.objects.only("costo").get(pk=g["product_id"]), "costo", None)
        articulos = list(Articulo.objects.select_related("product_id").filter(articulo_id__in=ids))
        baja = BajaStock.objects.create(usuario=request.user, local=local or "")
        movs = [
            MovimientoStock(
                tipo=MovimientoStock.Tipo.BAJA,
                local=local,
                usuario=request.user,
                articulo=a,
                producto=a.product_id,
                sku=a.sku,
                barcode=a.barcode,
                talle=a.talle,
                color=a.color,
                cantidad=1,
                costo_unitario=costo,
                baja=baja,
                precio_unitario=None,
                profit_unitario=Decimal("0.00"),
                nota="Baja por lote",
            )
            for a in articulos
        ]
        MovimientoStock.objects.bulk_create(movs)

        messages.success(request, f"Baja aplicada a {len(ids)} artículo(s).")
        return redirect("inventory:articulo_list")

class ArticuloListView(LoginRequiredMixin, ListView):
    model = Articulo
    template_name = "inventory/articulo/articulo_list.html"
    context_object_name = "articulos"
    paginate_by = 20

    def get_queryset(self):
        qs = _articulos_visibles_qs(self.request).select_related(
            "product_id",
            "product_id__marca",
            "local",
        )

        estado = (self.request.GET.get("estado") or "DISP").strip().upper()
        if estado in ["DISP", "VEND", "BAJA"]:
            qs = qs.filter(estado=estado)

        scan = (self.request.GET.get("scan") or "").strip()
        if scan:
            return qs.filter(barcode=scan, estado="DISP").order_by("created_at", "articulo_id")

        q = (self.request.GET.get("q") or "").strip()
        field = (self.request.GET.get("field") or "all").strip().lower()

        allowed_fields = ["all", "sku", "barcode", "producto", "marca", "color", "talle", "id"]
        if field not in allowed_fields:
            field = "all"

        if (not self.request.user.is_staff) and (not q):
            return qs.none()

        if q:
            if field == "all":
                filt = (
                    Q(sku__icontains=q) |
                    Q(color__icontains=q) |
                    Q(barcode__icontains=q) |
                    Q(product_id__nombre__icontains=q) |
                    Q(product_id__marca__nombre__icontains=q)
                )
                if q.isdigit():
                    n = int(q)
                    filt |= Q(talle=n) | Q(articulo_id=n)
                qs = qs.filter(filt)

            elif field == "sku":
                qs = qs.filter(sku__icontains=q)
            elif field == "barcode":
                qs = qs.filter(barcode__icontains=q)
            elif field == "producto":
                qs = qs.filter(product_id__nombre__icontains=q)
            elif field == "marca":
                qs = qs.filter(product_id__marca__nombre__icontains=q)
            elif field == "color":
                qs = qs.filter(color__icontains=q)
            elif field == "talle":
                qs = qs.filter(talle=int(q)) if q.isdigit() else qs.none()
            elif field == "id":
                qs = qs.filter(articulo_id=int(q)) if q.isdigit() else qs.none()

        mode = (self.request.GET.get("mode") or "qty").strip().lower()
        if mode not in ["qty", "unit"]:
            mode = "qty"

        if mode == "unit":
            return qs.order_by("created_at", "articulo_id")

        values_fields = [
            "sku",
            "barcode",
            "talle",
            "color",
            "estado",
            "product_id",
            "product_id__nombre",
            "product_id__marca__nombre",
        ]

        if _should_show_all_locals(self.request):
            values_fields.extend([
                "local_id",
                "local__nombre",
            ])

        return (
            qs.values(*values_fields)
            .annotate(
                qty=Count("articulo_id"),
                last_created=Max("created_at"),
            )
            .order_by("-last_created", "barcode", "talle", "color")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        local_activo = _get_local_activo(self.request)

        ctx["locales"] = Local.objects.all().order_by("nombre")
        ctx["local_activo"] = local_activo
        ctx["local_activo_nombre"] = local_activo.nombre if local_activo else "—"
        ctx["all_locals_active"] = _should_show_all_locals(self.request)
        ctx["scan"] = (self.request.GET.get("scan") or "").strip()
        ctx["q"] = (self.request.GET.get("q") or "").strip()

        field = (self.request.GET.get("field") or "all").strip().lower()
        if field not in ["all", "sku", "producto", "marca", "color", "talle", "id", "barcode"]:
            field = "all"
        ctx["field"] = field

        ctx["estado"] = (self.request.GET.get("estado") or "DISP").upper()

        mode = (self.request.GET.get("mode") or "qty").strip().lower()
        if ctx["scan"]:
            mode = "unit"
        if mode not in ["qty", "unit"]:
            mode = "qty"
        ctx["mode"] = mode

        ctx["auto_open_first"] = bool(ctx["scan"] and ctx["mode"] == "unit" and ctx["articulos"])
        return ctx


class ArticuloCreateView(LoginRequiredMixin, FormView):
    template_name = "inventory/articulo/articulo_create.html"
    form_class = ArticuloCreateForm
    success_url = reverse_lazy("inventory:articulo_list")
    
    def get_initial(self):
        
        initial = super().get_initial()
        local = _get_local_activo(self.request)
        barcode = (self.request.GET.get("barcode") or "").strip()

        if local and barcode:
            art = (
                Articulo.objects
                .select_related("product_id")
                .filter(local=local, barcode=barcode)
                .order_by("-articulo_id")
                .first()
            )
            if art:
                initial.update({
                    "barcode": barcode,
                    "product_id": art.product_id,
                    "talle": art.talle,
                    "color": art.color,
                    "sku_preview": build_sku(art.product_id, art.color, art.talle),
                })
        return initial
    
    @transaction.atomic
    def form_valid(self, form):
        producto = form.cleaned_data['product_id']
        barcode = form.cleaned_data['barcode']
        talle = form.cleaned_data['talle']
        color = (form.cleaned_data['color'] or "").strip()
        cantidad = form.cleaned_data['cantidad']
        local = _get_local_activo(self.request)
        referencia = (form.cleaned_data.get('referencia') or "").strip()
        costo_unitario = getattr(producto, "costo", None)
        sku = build_sku(producto, color, talle)

        existente = (
            Articulo.objects
            .select_for_update()
            .filter(local=local, barcode=barcode)
            .order_by("-articulo_id")
            .first()
        )

        if existente:
            if (
                existente.product_id_id != producto.pk
                or existente.talle != talle
                or (existente.color or "").strip().lower() != color.lower()
            ):
                form.add_error("barcode", "Ese barcode ya existe y pertenece a otro artículo.")
                return self.form_invalid(form)

        ingreso = Ingreso.objects.create(
            usuario=self.request.user,
            local=local,
            referencia=referencia,
            nota="Ingreso por carga de artículos"
        )

        total_linea = (costo_unitario * Decimal(cantidad)) if costo_unitario is not None else None
        item = IngresoItem.objects.create(
            ingreso=ingreso,
            producto=producto,
            sku=sku,
            barcode=barcode,
            talle=talle,
            color=color,
            cantidad=cantidad,
            costo_unitario=costo_unitario,
            total_linea=total_linea
        )

        articulos = [
            Articulo(
                product_id=producto,
                sku=sku,
                barcode=barcode,
                estado=Articulo.Estado.DISPONIBLE,
                talle=talle,
                color=color,
                local=local,
                ingreso_item=item
            )
            for _ in range(cantidad)
        ]
        Articulo.objects.bulk_create(articulos)

        creados = list(
            Articulo.objects
            .filter(ingreso_item=item, local=local)
            .order_by("articulo_id")[:cantidad]
        )

        movs = [
            MovimientoStock(
                tipo=MovimientoStock.Tipo.INGRESO,
                local=local,
                usuario=self.request.user,
                articulo=a,
                producto=producto,
                sku=sku,
                barcode=barcode,
                talle=talle,
                color=color,
                cantidad=1,
                costo_unitario=costo_unitario,
                precio_unitario=None,
                profit_unitario=Decimal("0.00"),
                ingreso=ingreso,
                venta=None,
                nota=f"Ingreso #{ingreso.ingreso_id}",
            )
            for a in creados
        ]
        MovimientoStock.objects.bulk_create(movs)

        messages.success(
            self.request,
            f"Ingreso #{ingreso.ingreso_id} registrado: {cantidad} unidad(es) de {producto}."
        )
        return super().form_valid(form)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["local"] = _get_local_activo(self.request)
        return kwargs


class ArticuloLookupByBarcodeView(LoginRequiredMixin, View):
    def get(self, request):
        barcode = (request.GET.get("barcode") or "").strip()
        if not barcode:
            return JsonResponse({"found": False}, status=400)

        qs = _articulos_visibles_qs(request)

        art = (
            qs.filter(barcode=barcode)
              .order_by("-articulo_id")
              .first()
        )

        if not art:
            return JsonResponse({"found": False})

        return JsonResponse({
            "found": True,
            "product_id": art.product_id.product_id,
            "talle": art.talle,
            "color": art.color,
        })


class ArticuloImportXlsxView(FormView):
    template_name = "inventory/articulo/articulo_import_xlsx.html"
    form_class = ArticuloImportXlsxForm
    success_url = reverse_lazy("inventory:articulo_list")

    @transaction.atomic
    def form_valid(self, form):
        local = _get_local_activo(self.request)
        if not local:
            messages.error(self.request, "No hay un local seleccionado.")
            return redirect("inventory:articulo_list")

        f = form.cleaned_data["file"]

        try:
            wb = load_workbook(filename=f, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(self.request, "Error al leer documento Excel. Asegurate que sea .xlsx válido.")
            return self.form_invalid(form)

        # Leer encabezados
        header = []
        for cell in ws[1]:
            header.append(_norm(cell.value).lower())

        required = ["barcode", "producto_nombre", "talle", "color", "cantidad"]
        missing = [c for c in required if c not in header]
        if missing:
            messages.error(self.request, f"Faltan columnas: {', '.join(missing)}")
            return self.form_invalid(form)

        idx = {name: header.index(name) for name in header if name}

        creados_total = 0
        filas_ok = 0
        errores = []

        # Procesar filas
        ingreso = Ingreso.objects.create(
                usuario=self.request.user,
                local=local,
                referencia="IMPORT_XLSX",
                nota="Ingreso por importación Excel"
            )
        for row_num in range(2, ws.max_row + 1):
            row = [ws.cell(row=row_num, column=col).value for col in range(1, ws.max_column + 1)]

            barcode = _norm(row[idx["barcode"]])
            producto_nombre = _norm(row[idx["producto_nombre"]])
            talle = _to_int(row[idx["talle"]])
            color = _norm(row[idx["color"]])
            cantidad = _to_int(row[idx["cantidad"]], 0)
            referencia = _norm(row[idx["referencia"]]) if "referencia" in idx else ""

            if not barcode and not producto_nombre and not talle and not color and not cantidad:
                continue

            if not barcode:
                errores.append(f"Fila {row_num}: barcode vacío")
                continue
            if not producto_nombre:
                errores.append(f"Fila {row_num}: producto vacío")
                continue
            producto = Producto.objects.filter(nombre=producto_nombre).first()
            if not producto:
                errores.append(f"Fila {row_num}: no existe producto con nombre '{producto_nombre}'")
                continue
            if cantidad <= 0:
                errores.append(f"Fila {row_num}: cantidad inválida")
                continue
            
            conflicto = _barcode_conflict(local, barcode, producto, talle, color)
            if conflicto:
                errores.append(
                    f"Fila {row_num}: el barcode '{barcode}' ya existe y está asociado a "
                    f"'{conflicto.product_id}' / Color: '{conflicto.color}' / Talle: '{conflicto.talle}'"
                )
                continue
            
            # ---- MISMA LÓGICA QUE TU form_valid ----
            costo_unitario = getattr(producto, "costo", None)
            sku = build_sku(producto, color, talle)

            

            total_linea = (costo_unitario * Decimal(cantidad)) if costo_unitario is not None else None
            item = IngresoItem.objects.create(
                ingreso=ingreso,
                producto=producto,
                sku=sku,
                barcode=barcode,
                talle=talle,
                color=color,
                cantidad=cantidad,
                costo_unitario=costo_unitario,
                total_linea=total_linea
            )

            articulos = [
                Articulo(
                    product_id=producto,
                    sku=sku,
                    barcode=barcode,
                    estado=Articulo.Estado.DISPONIBLE,
                    talle=talle,
                    color=color,
                    local=local,
                    ingreso_item=item
                )
                for _ in range(cantidad)
            ]
            Articulo.objects.bulk_create(articulos)

            creados = list(
                Articulo.objects
                .filter(ingreso_item=item, local=local)
                .order_by("articulo_id")[:cantidad]
            )

            movs = [
                MovimientoStock(
                    tipo=MovimientoStock.Tipo.INGRESO,
                    local=local,
                    usuario=self.request.user,
                    articulo=a,
                    producto=producto,
                    sku=sku,
                    barcode=barcode,
                    talle=talle,
                    color=color,
                    cantidad=1,
                    costo_unitario=costo_unitario,
                    precio_unitario=None,
                    profit_unitario=Decimal("0.00"),
                    ingreso=ingreso,
                    venta=None,
                    nota=f"Ingreso #{ingreso.ingreso_id}",
                )
                for a in creados
            ]
            MovimientoStock.objects.bulk_create(movs)

            filas_ok += 1
            creados_total += cantidad

        # Feedback
        if filas_ok:
            messages.success(self.request, f"Importación OK: {filas_ok} fila(s), {creados_total} artículo(s).")
        if errores:
            # no lo hagas eterno; mostramos las primeras 10
            preview = "\n".join(errores[:10])
            messages.warning(self.request, f"Hubo errores en algunas filas:\n{preview}")

        return super().form_valid(form)
